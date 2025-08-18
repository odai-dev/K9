"""
Unified Attendance Matrix Export Functions
PDF, Excel, and CSV export functionality for unified attendance matrix
"""

from datetime import datetime, date
from typing import Dict, Any
import os
import logging
import csv

from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.units import inch
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

from models import User
from unified_services import get_unified_matrix
from utils_pdf_rtl import rtl  # Arabic RTL text handling

logger = logging.getLogger(__name__)


def export_pdf(filters: dict, user: User) -> str:
    """
    Export unified attendance matrix to PDF with Arabic RTL support
    
    Args:
        filters: Filter parameters for the matrix
        user: Current user object
        
    Returns:
        str: Relative path to the generated PDF file
    """
    # Get matrix data
    matrix_data = get_unified_matrix(filters, user)
    
    # Create filename with date range
    date_from = filters['date_from'].replace('-', '')
    date_to = filters['date_to'].replace('-', '')
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f"unified_matrix_{date_from}-{date_to}_{timestamp}.pdf"
    
    # Ensure output directory exists
    output_dir = os.path.join('uploads', 'reports', '2025', '08')
    os.makedirs(output_dir, exist_ok=True)
    file_path = os.path.join(output_dir, filename)
    
    # Register Arabic font if available
    try:
        arabic_font_path = "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf"
        if os.path.exists(arabic_font_path):
            pdfmetrics.registerFont(TTFont('DejaVuSans', arabic_font_path))
            print(f"Successfully registered DejaVu Sans font from: {arabic_font_path}")
        else:
            print("Arabic font not found, using default")
    except Exception as e:
        print(f"Font registration error: {e}")
    
    # Create PDF document in landscape mode
    doc = SimpleDocTemplate(
        file_path,
        pagesize=landscape(A4),
        rightMargin=inch*0.5,
        leftMargin=inch*0.5,
        topMargin=inch*0.7,
        bottomMargin=inch*0.5
    )
    
    # Build content
    story = []
    
    # Add title and date range
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Title'],
        fontName='DejaVuSans' if os.path.exists("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf") else 'Helvetica-Bold',
        fontSize=16,
        alignment=2,  # Right alignment for RTL
        spaceAfter=12
    )
    
    title_text = rtl("المصفوفة الموحدة للحضور")
    story.append(Paragraph(title_text, title_style))
    
    # Date range
    date_from_formatted = datetime.strptime(filters['date_from'], '%Y-%m-%d').strftime('%d/%m/%Y')
    date_to_formatted = datetime.strptime(filters['date_to'], '%Y-%m-%d').strftime('%d/%m/%Y')
    date_range_text = rtl(f"الفترة: {date_from_formatted} - {date_to_formatted}")
    story.append(Paragraph(date_range_text, title_style))
    story.append(Spacer(1, 12))
    
    # Build table data
    table_data = _build_pdf_table_data(matrix_data, filters.get('include_dogs', False))
    
    if not table_data:
        no_data_text = rtl("لا توجد بيانات للعرض")
        story.append(Paragraph(no_data_text, styles['Normal']))
    else:
        # Create table with RTL column order (newest dates on left)
        table = Table(table_data, repeatRows=1)
        
        # Apply table styling
        table_style = _get_pdf_table_style(len(table_data[0]) if table_data else 0)
        table.setStyle(table_style)
        
        story.append(table)
    
    # Build PDF
    doc.build(story)
    
    logger.info(f"Unified matrix PDF exported: {file_path}")
    
    # Return relative path
    return file_path.replace('uploads/', '')


def export_excel(filters: dict, user: User) -> str:
    """
    Export unified attendance matrix to Excel with Arabic support
    
    Args:
        filters: Filter parameters for the matrix
        user: Current user object
        
    Returns:
        str: Relative path to the generated Excel file
    """
    if not OPENPYXL_AVAILABLE:
        raise ImportError("openpyxl is required for Excel export")
    
    # Get matrix data
    matrix_data = get_unified_matrix(filters, user)
    
    # Create filename
    date_from = filters['date_from'].replace('-', '')
    date_to = filters['date_to'].replace('-', '')
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f"unified_matrix_{date_from}-{date_to}_{timestamp}.xlsx"
    
    # Ensure output directory exists
    output_dir = os.path.join('uploads', 'reports', '2025', '08')
    os.makedirs(output_dir, exist_ok=True)
    file_path = os.path.join(output_dir, filename)
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Unified Attendance Matrix"
    
    # Set up styles
    header_font = Font(bold=True, size=12)
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    center_alignment = Alignment(horizontal="center", vertical="center")
    
    # Add title and date range
    ws.merge_cells('A1:Z1')
    ws['A1'] = "المصفوفة الموحدة للحضور"
    ws['A1'].font = Font(bold=True, size=16)
    ws['A1'].alignment = center_alignment
    
    date_range = f"الفترة: {filters['date_from']} - {filters['date_to']}"
    ws.merge_cells('A2:Z2')
    ws['A2'] = date_range
    ws['A2'].font = Font(bold=True, size=12)
    ws['A2'].alignment = center_alignment
    
    # Build and write table data
    table_data = _build_excel_table_data(matrix_data, filters.get('include_dogs', False))
    
    if table_data:
        start_row = 4
        for row_idx, row_data in enumerate(table_data):
            for col_idx, cell_value in enumerate(row_data):
                cell = ws.cell(row=start_row + row_idx, column=col_idx + 1, value=cell_value)
                
                # Style header row
                if row_idx == 0:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = center_alignment
        
        # Freeze header row and first column
        ws.freeze_panes = ws.cell(row=start_row + 1, column=2)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 20)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save workbook
    wb.save(file_path)
    
    logger.info(f"Unified matrix Excel exported: {file_path}")
    
    # Return relative path
    return file_path.replace('uploads/', '')


def export_csv(filters: dict, user: User) -> str:
    """
    Export unified attendance matrix to CSV
    
    Args:
        filters: Filter parameters for the matrix
        user: Current user object
        
    Returns:
        str: Relative path to the generated CSV file
    """
    # Get matrix data
    matrix_data = get_unified_matrix(filters, user)
    
    # Create filename
    date_from = filters['date_from'].replace('-', '')
    date_to = filters['date_to'].replace('-', '')
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f"unified_matrix_{date_from}-{date_to}_{timestamp}.csv"
    
    # Ensure output directory exists
    output_dir = os.path.join('uploads', 'reports', '2025', '08')
    os.makedirs(output_dir, exist_ok=True)
    file_path = os.path.join(output_dir, filename)
    
    # Build table data
    table_data = _build_csv_table_data(matrix_data, filters.get('include_dogs', False))
    
    # Write CSV file
    with open(file_path, 'w', newline='', encoding='utf-8-sig') as csvfile:
        writer = csv.writer(csvfile)
        for row in table_data:
            writer.writerow(row)
    
    logger.info(f"Unified matrix CSV exported: {file_path}")
    
    # Return relative path
    return file_path.replace('uploads/', '')


def _build_pdf_table_data(matrix_data: dict, include_dogs: bool) -> list:
    """Build table data for PDF export with RTL column order"""
    if not matrix_data['rows']:
        return []
    
    # Build header row
    header = []
    if include_dogs:
        header.extend([rtl("اسم الموظف"), rtl("اسم الكلب")])
    else:
        header.append(rtl("اسم الموظف"))
    
    # Add date columns in REVERSE order for RTL (newest dates on left)
    days = matrix_data['days']
    for day in reversed(days):
        day_formatted = datetime.strptime(day, '%Y-%m-%d').strftime('%d/%m')
        header.append(rtl(day_formatted))
    
    table_data = [header]
    
    # Status labels for Arabic display
    status_labels = {
        'PRESENT': 'حاضر',
        'ABSENT': 'غائب',
        'LATE': 'متأخر',
        'SICK': 'مرضية',
        'LEAVE': 'إجازة',
        'REMOTE': 'عن بُعد',
        'OVERTIME': 'عمل إضافي'
    }
    
    # Build data rows
    for row in matrix_data['rows']:
        data_row = []
        
        # Employee name (and dog name if included)
        if include_dogs:
            data_row.extend([rtl(row['employee_name']), rtl(row.get('dog_name', ''))])
        else:
            data_row.append(rtl(row['employee_name']))
        
        # Add cells in REVERSE order to match RTL header
        cells = row['cells']
        for cell in reversed(cells):
            if cell['status'] and not cell['project_controlled']:
                status_text = status_labels.get(cell['status'], cell['status'])
                if cell['check_in_time'] and cell['check_out_time']:
                    cell_text = f"{status_text}\n{cell['check_in_time']}-{cell['check_out_time']}"
                else:
                    cell_text = status_text
                data_row.append(rtl(cell_text))
            else:
                data_row.append('')  # Empty for project-controlled or null status
        
        table_data.append(data_row)
    
    return table_data


def _build_excel_table_data(matrix_data: dict, include_dogs: bool) -> list:
    """Build table data for Excel export with RTL column order"""
    if not matrix_data['rows']:
        return []
    
    # Build header row
    header = []
    if include_dogs:
        header.extend(["اسم الموظف", "اسم الكلب"])
    else:
        header.append("اسم الموظف")
    
    # Add date columns in REVERSE order for RTL
    days = matrix_data['days']
    for day in reversed(days):
        day_formatted = datetime.strptime(day, '%Y-%m-%d').strftime('%d/%m')
        header.append(day_formatted)
    
    table_data = [header]
    
    # Status labels for Arabic display
    status_labels = {
        'PRESENT': 'حاضر',
        'ABSENT': 'غائب',
        'LATE': 'متأخر',
        'SICK': 'مرضية',
        'LEAVE': 'إجازة',
        'REMOTE': 'عن بُعد',
        'OVERTIME': 'عمل إضافي'
    }
    
    # Build data rows
    for row in matrix_data['rows']:
        data_row = []
        
        # Employee name (and dog name if included)
        if include_dogs:
            data_row.extend([row['employee_name'], row.get('dog_name', '')])
        else:
            data_row.append(row['employee_name'])
        
        # Add cells in REVERSE order to match RTL header
        cells = row['cells']
        for cell in reversed(cells):
            if cell['status'] and not cell['project_controlled']:
                status_text = status_labels.get(cell['status'], cell['status'])
                data_row.append(status_text)
            else:
                data_row.append('')  # Empty for project-controlled or null status
        
        table_data.append(data_row)
    
    return table_data


def _build_csv_table_data(matrix_data: dict, include_dogs: bool) -> list:
    """Build table data for CSV export with RTL column order"""
    # CSV uses same structure as Excel
    return _build_excel_table_data(matrix_data, include_dogs)


def _get_pdf_table_style(num_cols: int) -> TableStyle:
    """Get table style for PDF export"""
    style_commands = [
        # Header row styling
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'DejaVuSans' if os.path.exists("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf") else 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        
        # Data rows styling
        ('FONTNAME', (0, 1), (-1, -1), 'DejaVuSans' if os.path.exists("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf") else 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        
        # Alternating row colors
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f0f0f0')]),
    ]
    
    return TableStyle(style_commands)