from flask import current_app
"""
Report Export Service
Handles PDF and Excel export for all report types

Updated to use Minimal Elegant PDF design template
"""
import os
from io import BytesIO
from datetime import datetime
from typing import Optional, BinaryIO

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.lib.enums import TA_RIGHT, TA_CENTER
from reportlab.pdfgen import canvas

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from k9.models.models_handler_daily import HandlerReport, ShiftReport, ReportStatus
from k9.models.models import VeterinaryVisit, BreedingTrainingActivity, CaretakerDailyLog
from k9.utils.pdf_minimal_elegant import (
    create_minimal_header, 
    create_info_section,
    create_text_section,
    create_data_table,
    get_minimal_table_style,
    get_minimal_styles,
    add_page_number,
    MinimalColors
)
from k9.utils.utils_pdf_rtl import rtl, register_arabic_fonts, get_arabic_font_name, format_pdf_text


class ReportExportService:
    """Service for exporting reports to PDF and Excel"""
    
    @staticmethod
    def _get_table_style_base(header_color='#c5cae9'):
        """Get base table style with proper font fallback"""
        register_arabic_fonts()
        font_name = get_arabic_font_name()
        return TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor(header_color)),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, -1), font_name),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ])
    
    @staticmethod
    def _get_arabic_styles():
        """Get paragraph styles configured for Arabic text"""
        styles = getSampleStyleSheet()
        register_arabic_fonts()
        font_name = get_arabic_font_name()
        
        # Title style
        title_style = ParagraphStyle(
            'ArabicTitle',
            parent=styles['Title'],
            fontName=font_name,
            fontSize=18,
            alignment=TA_CENTER,
            spaceAfter=12,
            textColor=colors.HexColor('#1a237e')
        )
        
        # Heading style
        heading_style = ParagraphStyle(
            'ArabicHeading',
            parent=styles['Heading1'],
            fontName=font_name,
            fontSize=14,
            alignment=TA_RIGHT,
            spaceAfter=8,
            textColor=colors.HexColor('#283593')
        )
        
        # Normal text style
        normal_style = ParagraphStyle(
            'ArabicNormal',
            parent=styles['Normal'],
            fontName=font_name,
            fontSize=10,
            alignment=TA_RIGHT,
            spaceAfter=6
        )
        
        return {
            'title': title_style,
            'heading': heading_style,
            'normal': normal_style
        }
    
    @staticmethod
    def export_handler_report_to_pdf(report_id: str) -> Optional[BytesIO]:
        """
        Export HandlerReport to PDF
        
        Args:
            report_id: The UUID of the report
            
        Returns:
            BytesIO object containing the PDF, or None if report not found
        """
        report = HandlerReport.query.get(report_id)
        if not report:
            return None
        
        buffer = BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            rightMargin=2*cm,
            leftMargin=2*cm,
            topMargin=2*cm,
            bottomMargin=2*cm
        )
        
        styles = ReportExportService._get_arabic_styles()
        register_arabic_fonts()
        font_name = get_arabic_font_name()
        story = []
        
        # Title
        title = Paragraph(rtl("تقرير السائس اليومي"), styles['title'])
        story.append(title)
        story.append(Spacer(1, 0.5*cm))
        
        # Report metadata
        metadata_data = [
            [rtl('التاريخ'), report.date.strftime('%Y-%m-%d')],
            [rtl('نوع التقرير'), rtl('يومي شامل' if report.report_type.value == 'DAILY' else 'وردية')],
            [rtl('السائس'), format_pdf_text(report.handler.username if report.handler else 'غير محدد')],
            [rtl('الكلب'), format_pdf_text(report.dog.name if report.dog else 'غير محدد')],
            [rtl('المشروع'), format_pdf_text(report.project.name if report.project else 'غير محدد')],
            [rtl('الموقع'), format_pdf_text(report.location or 'غير محدد')],
            [rtl('الحالة'), rtl(ReportExportService._get_status_arabic(report.status.value))],
        ]
        
        if report.submitted_at:
            metadata_data.append([rtl('تاريخ الإرسال'), report.submitted_at.strftime('%Y-%m-%d %H:%M')])
        
        if report.reviewer:
            metadata_data.append([rtl('المراجع'), format_pdf_text(report.reviewer.username)])
            if report.reviewed_at:
                metadata_data.append([rtl('تاريخ المراجعة'), report.reviewed_at.strftime('%Y-%m-%d %H:%M')])
        
        if report.review_notes:
            metadata_data.append([rtl('ملاحظات المراجعة'), format_pdf_text(report.review_notes)])
        
        metadata_table = Table(metadata_data, colWidths=[5*cm, 10*cm])
        metadata_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#e8eaf6')),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, -1), font_name),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ]))
        
        story.append(metadata_table)
        story.append(Spacer(1, 0.7*cm))
        
        # Health section
        if report.health:
            story.append(Paragraph(rtl("الفحص الصحي"), styles['heading']))
            health_data = [[rtl('الجزء'), rtl('الحالة'), rtl('الملاحظات')]]
            
            health_fields = [
                ('eyes', 'العيون'),
                ('nose', 'الأنف'),
                ('ears', 'الأذنين'),
                ('mouth', 'الفم'),
                ('teeth', 'الأسنان'),
                ('gums', 'اللثة'),
                ('front_limbs', 'الأطراف الأمامية'),
                ('back_limbs', 'الأطراف الخلفية'),
                ('hair', 'الشعر'),
                ('tail', 'الذيل'),
                ('rear', 'الخلفية'),
            ]
            
            for field, label in health_fields:
                status = getattr(report.health, f'{field}_status', None)
                notes = getattr(report.health, f'{field}_notes', None)
                if status:
                    health_data.append([
                        rtl(label),
                        rtl(ReportExportService._get_health_status_arabic(status.value)),
                        format_pdf_text(notes or '')
                    ])
            
            if len(health_data) > 1:
                health_table = Table(health_data, colWidths=[4*cm, 3*cm, 8*cm])
                health_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#c5cae9')),
                    ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
                    ('ALIGN', (0, 0), (-1, -1), 'RIGHT'),
                    ('FONTNAME', (0, 0), (-1, -1), font_name),
                    ('FONTSIZE', (0, 0), (-1, -1), 9),
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                    ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ]))
                story.append(health_table)
                story.append(Spacer(1, 0.5*cm))
        
        # Training sessions
        if report.training_sessions:
            story.append(Paragraph(rtl("جلسات التدريب"), styles['heading']))
            training_data = [[rtl('النوع'), rtl('الوصف'), rtl('من'), rtl('إلى'), rtl('ملاحظات')]]
            
            for session in report.training_sessions:
                training_data.append([
                    format_pdf_text(session.training_type.value if session.training_type else ''),
                    format_pdf_text(session.description or ''),
                    session.time_from.strftime('%H:%M') if session.time_from else '',
                    session.time_to.strftime('%H:%M') if session.time_to else '',
                    format_pdf_text(session.notes or '')
                ])
            
            training_table = Table(training_data, colWidths=[3*cm, 4*cm, 2*cm, 2*cm, 4*cm])
            training_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#c5cae9')),
                ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
                ('ALIGN', (0, 0), (-1, -1), 'RIGHT'),
                ('FONTNAME', (0, 0), (-1, -1), font_name),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ]))
            story.append(training_table)
            story.append(Spacer(1, 0.5*cm))
        
        # Care section
        if report.care:
            story.append(Paragraph(rtl("العناية"), styles['heading']))
            care_data = []
            
            if report.care.food_amount:
                care_data.append([rtl('كمية الطعام'), format_pdf_text(report.care.food_amount)])
            if report.care.food_type:
                care_data.append([rtl('نوع الطعام'), format_pdf_text(report.care.food_type)])
            if report.care.supplements:
                care_data.append([rtl('المكملات'), format_pdf_text(report.care.supplements)])
            if report.care.water_amount:
                care_data.append([rtl('كمية الماء'), format_pdf_text(report.care.water_amount)])
            
            care_data.append([rtl('التمشيط'), rtl('نعم' if report.care.grooming_done else 'لا')])
            care_data.append([rtl('الغسل'), rtl('نعم' if report.care.washing_done else 'لا')])
            
            if report.care.excretion_location:
                care_data.append([rtl('مكان التبرز'), format_pdf_text(report.care.excretion_location)])
            if report.care.stool_color:
                care_data.append([rtl('لون البراز'), format_pdf_text(report.care.stool_color.value)])
            if report.care.stool_shape:
                care_data.append([rtl('شكل البراز'), format_pdf_text(report.care.stool_shape.value)])
            
            if care_data:
                care_table = Table(care_data, colWidths=[5*cm, 10*cm])
                care_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#e8eaf6')),
                    ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
                    ('ALIGN', (0, 0), (-1, -1), 'RIGHT'),
                    ('FONTNAME', (0, 0), (-1, -1), font_name),
                    ('FONTSIZE', (0, 0), (-1, -1), 10),
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                    ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ]))
                story.append(care_table)
                story.append(Spacer(1, 0.5*cm))
        
        # Behavior section
        if report.behavior:
            story.append(Paragraph(rtl("السلوك"), styles['heading']))
            behavior_data = []
            
            if report.behavior.good_behavior_notes:
                behavior_data.append([rtl('السلوك الجيد'), format_pdf_text(report.behavior.good_behavior_notes)])
            if report.behavior.bad_behavior_notes:
                behavior_data.append([rtl('السلوك السيئ'), format_pdf_text(report.behavior.bad_behavior_notes)])
            
            if behavior_data:
                behavior_table = Table(behavior_data, colWidths=[5*cm, 10*cm])
                behavior_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#e8eaf6')),
                    ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
                    ('ALIGN', (0, 0), (-1, -1), 'RIGHT'),
                    ('FONTNAME', (0, 0), (-1, -1), font_name),
                    ('FONTSIZE', (0, 0), (-1, -1), 10),
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                    ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ]))
                story.append(behavior_table)
                story.append(Spacer(1, 0.5*cm))
        
        # Incidents section
        if report.incidents:
            story.append(Paragraph(rtl("الحوادث والاشتباهات"), styles['heading']))
            incident_data = [[rtl('النوع'), rtl('الوصف'), rtl('التاريخ والوقت'), rtl('الموقع')]]
            
            for incident in report.incidents:
                incident_data.append([
                    format_pdf_text(incident.incident_type.value if incident.incident_type else ''),
                    format_pdf_text(incident.description or ''),
                    incident.incident_datetime.strftime('%Y-%m-%d %H:%M') if incident.incident_datetime else '',
                    format_pdf_text(incident.location or '')
                ])
            
            incident_table = Table(incident_data, colWidths=[3*cm, 6*cm, 3.5*cm, 2.5*cm])
            incident_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#ffcdd2')),
                ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
                ('ALIGN', (0, 0), (-1, -1), 'RIGHT'),
                ('FONTNAME', (0, 0), (-1, -1), font_name),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ]))
            story.append(incident_table)
        
        # Build PDF
        doc.build(story)
        buffer.seek(0)
        return buffer
    
    @staticmethod
    def export_handler_report_to_excel(report_id: str) -> Optional[BytesIO]:
        """
        Export HandlerReport to Excel
        
        Args:
            report_id: The UUID of the report
            
        Returns:
            BytesIO object containing the Excel file, or None if report not found
        """
        report = HandlerReport.query.get(report_id)
        if not report:
            return None
        
        buffer = BytesIO()
        wb = Workbook()
        
        # Styles
        header_font = Font(bold=True, size=12, color='FFFFFF')
        header_fill = PatternFill(start_color='1a237e', end_color='1a237e', fill_type='solid')
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Main sheet
        ws = wb.active
        ws.title = 'معلومات التقرير'
        ws.right_to_left = True
        
        # Report metadata
        row = 1
        ws.merge_cells(f'A{row}:B{row}')
        ws[f'A{row}'] = 'تقرير السائس اليومي'
        ws[f'A{row}'].font = Font(bold=True, size=14)
        ws[f'A{row}'].alignment = Alignment(horizontal='center')
        row += 2
        
        metadata = [
            ('التاريخ', report.date.strftime('%Y-%m-%d')),
            ('نوع التقرير', 'يومي شامل' if report.report_type.value == 'DAILY' else 'وردية'),
            ('السائس', report.handler.username if report.handler else 'غير محدد'),
            ('الكلب', report.dog.name if report.dog else 'غير محدد'),
            ('المشروع', report.project.name if report.project else 'غير محدد'),
            ('الموقع', report.location or 'غير محدد'),
            ('الحالة', ReportExportService._get_status_arabic(report.status.value)),
        ]
        
        if report.submitted_at:
            metadata.append(('تاريخ الإرسال', report.submitted_at.strftime('%Y-%m-%d %H:%M')))
        
        if report.reviewer:
            metadata.append(('المراجع', report.reviewer.username))
            if report.reviewed_at:
                metadata.append(('تاريخ المراجعة', report.reviewed_at.strftime('%Y-%m-%d %H:%M')))
        
        if report.review_notes:
            metadata.append(('ملاحظات المراجعة', report.review_notes))
        
        for label, value in metadata:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'A{row}'].fill = PatternFill(start_color='e8eaf6', end_color='e8eaf6', fill_type='solid')
            ws[f'A{row}'].border = border
            ws[f'B{row}'].border = border
            row += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 40
        
        # Health sheet
        if report.health:
            ws_health = wb.create_sheet('الفحص الصحي')
            ws_health.right_to_left = True
            
            ws_health['A1'] = 'الجزء'
            ws_health['B1'] = 'الحالة'
            ws_health['C1'] = 'الملاحظات'
            
            for cell in ['A1', 'B1', 'C1']:
                ws_health[cell].font = header_font
                ws_health[cell].fill = header_fill
                ws_health[cell].border = border
            
            health_fields = [
                ('eyes', 'العيون'),
                ('nose', 'الأنف'),
                ('ears', 'الأذنين'),
                ('mouth', 'الفم'),
                ('teeth', 'الأسنان'),
                ('gums', 'اللثة'),
                ('front_limbs', 'الأطراف الأمامية'),
                ('back_limbs', 'الأطراف الخلفية'),
                ('hair', 'الشعر'),
                ('tail', 'الذيل'),
                ('rear', 'الخلفية'),
            ]
            
            row = 2
            for field, label in health_fields:
                status = getattr(report.health, f'{field}_status', None)
                notes = getattr(report.health, f'{field}_notes', None)
                if status:
                    ws_health[f'A{row}'] = label
                    ws_health[f'B{row}'] = ReportExportService._get_health_status_arabic(status.value)
                    ws_health[f'C{row}'] = notes or ''
                    for cell in [f'A{row}', f'B{row}', f'C{row}']:
                        ws_health[cell].border = border
                    row += 1
            
            ws_health.column_dimensions['A'].width = 20
            ws_health.column_dimensions['B'].width = 15
            ws_health.column_dimensions['C'].width = 50
        
        # Training sheet
        if report.training_sessions:
            ws_training = wb.create_sheet('التدريب')
            ws_training.right_to_left = True
            
            ws_training['A1'] = 'النوع'
            ws_training['B1'] = 'الوصف'
            ws_training['C1'] = 'من'
            ws_training['D1'] = 'إلى'
            ws_training['E1'] = 'ملاحظات'
            
            for cell in ['A1', 'B1', 'C1', 'D1', 'E1']:
                ws_training[cell].font = header_font
                ws_training[cell].fill = header_fill
                ws_training[cell].border = border
            
            row = 2
            for session in report.training_sessions:
                ws_training[f'A{row}'] = session.training_type.value if session.training_type else ''
                ws_training[f'B{row}'] = session.description or ''
                ws_training[f'C{row}'] = session.time_from.strftime('%H:%M') if session.time_from else ''
                ws_training[f'D{row}'] = session.time_to.strftime('%H:%M') if session.time_to else ''
                ws_training[f'E{row}'] = session.notes or ''
                for cell in [f'A{row}', f'B{row}', f'C{row}', f'D{row}', f'E{row}']:
                    ws_training[cell].border = border
                row += 1
            
            ws_training.column_dimensions['A'].width = 20
            ws_training.column_dimensions['B'].width = 30
            ws_training.column_dimensions['C'].width = 10
            ws_training.column_dimensions['D'].width = 10
            ws_training.column_dimensions['E'].width = 30
        
        # Care sheet
        if report.care:
            ws_care = wb.create_sheet('العناية')
            ws_care.right_to_left = True
            
            ws_care['A1'] = 'البند'
            ws_care['B1'] = 'القيمة'
            
            for cell in ['A1', 'B1']:
                ws_care[cell].font = header_font
                ws_care[cell].fill = header_fill
                ws_care[cell].border = border
            
            row = 2
            care_items = []
            
            if report.care.food_amount:
                care_items.append(('كمية الطعام', report.care.food_amount))
            if report.care.food_type:
                care_items.append(('نوع الطعام', report.care.food_type))
            if report.care.supplements:
                care_items.append(('المكملات', report.care.supplements))
            if report.care.water_amount:
                care_items.append(('كمية الماء', report.care.water_amount))
            
            care_items.append(('التمشيط', 'نعم' if report.care.grooming_done else 'لا'))
            care_items.append(('الغسل', 'نعم' if report.care.washing_done else 'لا'))
            
            if report.care.excretion_location:
                care_items.append(('مكان التبرز', report.care.excretion_location))
            if report.care.stool_color:
                care_items.append(('لون البراز', report.care.stool_color.value))
            if report.care.stool_shape:
                care_items.append(('شكل البراز', report.care.stool_shape.value))
            
            for label, value in care_items:
                ws_care[f'A{row}'] = label
                ws_care[f'B{row}'] = value
                ws_care[f'A{row}'].font = Font(bold=True)
                ws_care[f'A{row}'].fill = PatternFill(start_color='e8eaf6', end_color='e8eaf6', fill_type='solid')
                for cell in [f'A{row}', f'B{row}']:
                    ws_care[cell].border = border
                row += 1
            
            ws_care.column_dimensions['A'].width = 20
            ws_care.column_dimensions['B'].width = 40
        
        # Behavior sheet
        if report.behavior:
            ws_behavior = wb.create_sheet('السلوك')
            ws_behavior.right_to_left = True
            
            ws_behavior['A1'] = 'النوع'
            ws_behavior['B1'] = 'الملاحظات'
            
            for cell in ['A1', 'B1']:
                ws_behavior[cell].font = header_font
                ws_behavior[cell].fill = header_fill
                ws_behavior[cell].border = border
            
            row = 2
            if report.behavior.good_behavior_notes:
                ws_behavior[f'A{row}'] = 'السلوك الجيد'
                ws_behavior[f'B{row}'] = report.behavior.good_behavior_notes
                ws_behavior[f'A{row}'].font = Font(bold=True)
                ws_behavior[f'A{row}'].fill = PatternFill(start_color='c8e6c9', end_color='c8e6c9', fill_type='solid')
                for cell in [f'A{row}', f'B{row}']:
                    ws_behavior[cell].border = border
                row += 1
            
            if report.behavior.bad_behavior_notes:
                ws_behavior[f'A{row}'] = 'السلوك السيئ'
                ws_behavior[f'B{row}'] = report.behavior.bad_behavior_notes
                ws_behavior[f'A{row}'].font = Font(bold=True)
                ws_behavior[f'A{row}'].fill = PatternFill(start_color='ffcdd2', end_color='ffcdd2', fill_type='solid')
                for cell in [f'A{row}', f'B{row}']:
                    ws_behavior[cell].border = border
                row += 1
            
            ws_behavior.column_dimensions['A'].width = 20
            ws_behavior.column_dimensions['B'].width = 60
        
        # Incidents sheet
        if report.incidents:
            ws_incidents = wb.create_sheet('الحوادث')
            ws_incidents.right_to_left = True
            
            ws_incidents['A1'] = 'النوع'
            ws_incidents['B1'] = 'الوصف'
            ws_incidents['C1'] = 'التاريخ والوقت'
            ws_incidents['D1'] = 'الموقع'
            
            for cell in ['A1', 'B1', 'C1', 'D1']:
                ws_incidents[cell].font = header_font
                ws_incidents[cell].fill = PatternFill(start_color='d32f2f', end_color='d32f2f', fill_type='solid')
                ws_incidents[cell].border = border
            
            row = 2
            for incident in report.incidents:
                ws_incidents[f'A{row}'] = incident.incident_type.value if incident.incident_type else ''
                ws_incidents[f'B{row}'] = incident.description or ''
                ws_incidents[f'C{row}'] = incident.incident_datetime.strftime('%Y-%m-%d %H:%M') if incident.incident_datetime else ''
                ws_incidents[f'D{row}'] = incident.location or ''
                for cell in [f'A{row}', f'B{row}', f'C{row}', f'D{row}']:
                    ws_incidents[cell].border = border
                row += 1
            
            ws_incidents.column_dimensions['A'].width = 20
            ws_incidents.column_dimensions['B'].width = 40
            ws_incidents.column_dimensions['C'].width = 20
            ws_incidents.column_dimensions['D'].width = 20
        
        wb.save(buffer)
        buffer.seek(0)
        return buffer
    
    @staticmethod
    def _get_status_arabic(status: str) -> str:
        """Convert status to Arabic"""
        status_map = {
            'DRAFT': 'مسودة',
            'SUBMITTED': 'مرسل',
            'APPROVED': 'معتمد',
            'REJECTED': 'مرفوض',
            'APPROVED_BY_PM': 'معتمد من مدير المشروع',
            'FORWARDED_TO_ADMIN': 'محول للإدارة العامة',
            'REJECTED_BY_PM': 'مرفوض من مدير المشروع',
            'APPROVED_BY_ADMIN': 'معتمد من الإدارة العامة',
            'REJECTED_BY_ADMIN': 'مرفوض من الإدارة العامة',
        }
        return status_map.get(status, status)
    
    @staticmethod
    def _get_health_status_arabic(status: str) -> str:
        """Convert health status to Arabic"""
        status_map = {
            'NORMAL': 'طبيعي',
            'ABNORMAL': 'غير طبيعي',
            'NEEDS_VET': 'يحتاج بيطري',
        }
        return status_map.get(status, status)
    
    @staticmethod
    def export_shift_report_to_pdf(report_id: str) -> Optional[BytesIO]:
        """
        Export ShiftReport to PDF
        
        Args:
            report_id: The UUID of the shift report
            
        Returns:
            BytesIO object containing the PDF, or None if report not found
        """
        report = ShiftReport.query.get(report_id)
        if not report:
            return None
        
        buffer = BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            rightMargin=2*cm,
            leftMargin=2*cm,
            topMargin=2*cm,
            bottomMargin=2*cm
        )
        
        styles = ReportExportService._get_arabic_styles()
        register_arabic_fonts()
        font_name = get_arabic_font_name()
        story = []
        
        # Title
        title = Paragraph(rtl("تقرير الوردية"), styles['title'])
        story.append(title)
        story.append(Spacer(1, 0.5*cm))
        
        # Report metadata
        metadata_data = [
            [rtl('التاريخ'), report.date.strftime('%Y-%m-%d')],
            [rtl('السائس'), format_pdf_text(report.handler.username if report.handler else 'غير محدد')],
            [rtl('الكلب'), format_pdf_text(report.dog.name if report.dog else 'غير محدد')],
            [rtl('المشروع'), format_pdf_text(report.project.name if report.project else 'غير محدد')],
            [rtl('الموقع'), format_pdf_text(report.location or 'غير محدد')],
            [rtl('الحالة'), rtl(ReportExportService._get_status_arabic(report.status.value))],
        ]
        
        if report.submitted_at:
            metadata_data.append([rtl('تاريخ الإرسال'), report.submitted_at.strftime('%Y-%m-%d %H:%M')])
        
        if report.reviewer:
            metadata_data.append([rtl('المراجع'), format_pdf_text(report.reviewer.username)])
            if report.reviewed_at:
                metadata_data.append([rtl('تاريخ المراجعة'), report.reviewed_at.strftime('%Y-%m-%d %H:%M')])
        
        if report.review_notes:
            metadata_data.append([rtl('ملاحظات المراجعة'), format_pdf_text(report.review_notes)])
        
        metadata_table = Table(metadata_data, colWidths=[5*cm, 10*cm])
        metadata_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#e8eaf6')),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, -1), font_name),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ]))
        
        story.append(metadata_table)
        story.append(Spacer(1, 0.7*cm))
        
        # Health section
        if report.health:
            story.append(Paragraph(rtl("الفحص الصحي"), styles['heading']))
            health_data = [[rtl('الجزء'), rtl('الحالة'), rtl('الملاحظات')]]
            
            health_fields = [
                ('eyes', 'العيون'),
                ('nose', 'الأنف'),
                ('ears', 'الأذنين'),
                ('mouth', 'الفم'),
                ('teeth', 'الأسنان'),
                ('gums', 'اللثة'),
                ('front_limbs', 'الأطراف الأمامية'),
                ('back_limbs', 'الأطراف الخلفية'),
                ('hair', 'الشعر'),
                ('tail', 'الذيل'),
                ('rear', 'الخلفية'),
            ]
            
            for field, label in health_fields:
                status = getattr(report.health, f'{field}_status', None)
                notes = getattr(report.health, f'{field}_notes', None)
                if status:
                    health_data.append([
                        rtl(label),
                        rtl(ReportExportService._get_health_status_arabic(status.value)),
                        format_pdf_text(notes or '')
                    ])
            
            if len(health_data) > 1:
                health_table = Table(health_data, colWidths=[4*cm, 3*cm, 8*cm])
                health_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#c5cae9')),
                    ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
                    ('ALIGN', (0, 0), (-1, -1), 'RIGHT'),
                    ('FONTNAME', (0, 0), (-1, -1), font_name),
                    ('FONTSIZE', (0, 0), (-1, -1), 9),
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                    ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ]))
                story.append(health_table)
                story.append(Spacer(1, 0.5*cm))
        
        # Behavior section
        if report.behavior:
            story.append(Paragraph(rtl("السلوك"), styles['heading']))
            behavior_data = []
            
            if report.behavior.good_behavior_notes:
                behavior_data.append([rtl('السلوك الجيد'), format_pdf_text(report.behavior.good_behavior_notes)])
            if report.behavior.bad_behavior_notes:
                behavior_data.append([rtl('السلوك السيئ'), format_pdf_text(report.behavior.bad_behavior_notes)])
            
            if behavior_data:
                behavior_table = Table(behavior_data, colWidths=[5*cm, 10*cm])
                behavior_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#e8eaf6')),
                    ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
                    ('ALIGN', (0, 0), (-1, -1), 'RIGHT'),
                    ('FONTNAME', (0, 0), (-1, -1), font_name),
                    ('FONTSIZE', (0, 0), (-1, -1), 10),
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                    ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ]))
                story.append(behavior_table)
                story.append(Spacer(1, 0.5*cm))
        
        # Incidents section
        if report.incidents:
            story.append(Paragraph(rtl("الحوادث والاشتباهات"), styles['heading']))
            incident_data = [[rtl('النوع'), rtl('الوصف'), rtl('التاريخ والوقت'), rtl('الموقع')]]
            
            for incident in report.incidents:
                incident_data.append([
                    format_pdf_text(incident.incident_type.value if incident.incident_type else ''),
                    format_pdf_text(incident.description or ''),
                    incident.incident_datetime.strftime('%Y-%m-%d %H:%M') if incident.incident_datetime else '',
                    format_pdf_text(incident.location or '')
                ])
            
            incident_table = Table(incident_data, colWidths=[3*cm, 6*cm, 3.5*cm, 2.5*cm])
            incident_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#ffcdd2')),
                ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
                ('ALIGN', (0, 0), (-1, -1), 'RIGHT'),
                ('FONTNAME', (0, 0), (-1, -1), font_name),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ]))
            story.append(incident_table)
        
        # Build PDF
        doc.build(story)
        buffer.seek(0)
        return buffer
    
    @staticmethod
    def export_shift_report_to_excel(report_id: str) -> Optional[BytesIO]:
        """
        Export ShiftReport to Excel
        
        Args:
            report_id: The UUID of the shift report
            
        Returns:
            BytesIO object containing the Excel file, or None if report not found
        """
        report = ShiftReport.query.get(report_id)
        if not report:
            return None
        
        buffer = BytesIO()
        wb = Workbook()
        
        # Styles
        header_font = Font(bold=True, size=12, color='FFFFFF')
        header_fill = PatternFill(start_color='1a237e', end_color='1a237e', fill_type='solid')
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Main sheet
        ws = wb.active
        ws.title = 'معلومات التقرير'
        ws.right_to_left = True
        
        # Report metadata
        row = 1
        ws.merge_cells(f'A{row}:B{row}')
        ws[f'A{row}'] = 'تقرير الوردية'
        ws[f'A{row}'].font = Font(bold=True, size=14)
        ws[f'A{row}'].alignment = Alignment(horizontal='center')
        row += 2
        
        metadata = [
            ('التاريخ', report.date.strftime('%Y-%m-%d')),
            ('السائس', report.handler.username if report.handler else 'غير محدد'),
            ('الكلب', report.dog.name if report.dog else 'غير محدد'),
            ('المشروع', report.project.name if report.project else 'غير محدد'),
            ('الموقع', report.location or 'غير محدد'),
            ('الحالة', ReportExportService._get_status_arabic(report.status.value)),
        ]
        
        if report.submitted_at:
            metadata.append(('تاريخ الإرسال', report.submitted_at.strftime('%Y-%m-%d %H:%M')))
        
        if report.reviewer:
            metadata.append(('المراجع', report.reviewer.username))
            if report.reviewed_at:
                metadata.append(('تاريخ المراجعة', report.reviewed_at.strftime('%Y-%m-%d %H:%M')))
        
        if report.review_notes:
            metadata.append(('ملاحظات المراجعة', report.review_notes))
        
        for label, value in metadata:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'A{row}'].fill = PatternFill(start_color='e8eaf6', end_color='e8eaf6', fill_type='solid')
            ws[f'A{row}'].border = border
            ws[f'B{row}'].border = border
            row += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 40
        
        # Health sheet
        if report.health:
            ws_health = wb.create_sheet('الفحص الصحي')
            ws_health.right_to_left = True
            
            ws_health['A1'] = 'الجزء'
            ws_health['B1'] = 'الحالة'
            ws_health['C1'] = 'الملاحظات'
            
            for cell in ['A1', 'B1', 'C1']:
                ws_health[cell].font = header_font
                ws_health[cell].fill = header_fill
                ws_health[cell].border = border
            
            health_fields = [
                ('eyes', 'العيون'),
                ('nose', 'الأنف'),
                ('ears', 'الأذنين'),
                ('mouth', 'الفم'),
                ('teeth', 'الأسنان'),
                ('gums', 'اللثة'),
                ('front_limbs', 'الأطراف الأمامية'),
                ('back_limbs', 'الأطراف الخلفية'),
                ('hair', 'الشعر'),
                ('tail', 'الذيل'),
                ('rear', 'الخلفية'),
            ]
            
            row = 2
            for field, label in health_fields:
                status = getattr(report.health, f'{field}_status', None)
                notes = getattr(report.health, f'{field}_notes', None)
                if status:
                    ws_health[f'A{row}'] = label
                    ws_health[f'B{row}'] = ReportExportService._get_health_status_arabic(status.value)
                    ws_health[f'C{row}'] = notes or ''
                    for cell in [f'A{row}', f'B{row}', f'C{row}']:
                        ws_health[cell].border = border
                    row += 1
            
            ws_health.column_dimensions['A'].width = 20
            ws_health.column_dimensions['B'].width = 15
            ws_health.column_dimensions['C'].width = 50
        
        # Behavior sheet
        if report.behavior:
            ws_behavior = wb.create_sheet('السلوك')
            ws_behavior.right_to_left = True
            
            ws_behavior['A1'] = 'النوع'
            ws_behavior['B1'] = 'الملاحظات'
            
            for cell in ['A1', 'B1']:
                ws_behavior[cell].font = header_font
                ws_behavior[cell].fill = header_fill
                ws_behavior[cell].border = border
            
            row = 2
            if report.behavior.good_behavior_notes:
                ws_behavior[f'A{row}'] = 'السلوك الجيد'
                ws_behavior[f'B{row}'] = report.behavior.good_behavior_notes
                ws_behavior[f'A{row}'].font = Font(bold=True)
                ws_behavior[f'A{row}'].fill = PatternFill(start_color='c8e6c9', end_color='c8e6c9', fill_type='solid')
                for cell in [f'A{row}', f'B{row}']:
                    ws_behavior[cell].border = border
                row += 1
            
            if report.behavior.bad_behavior_notes:
                ws_behavior[f'A{row}'] = 'السلوك السيئ'
                ws_behavior[f'B{row}'] = report.behavior.bad_behavior_notes
                ws_behavior[f'A{row}'].font = Font(bold=True)
                ws_behavior[f'A{row}'].fill = PatternFill(start_color='ffcdd2', end_color='ffcdd2', fill_type='solid')
                for cell in [f'A{row}', f'B{row}']:
                    ws_behavior[cell].border = border
                row += 1
            
            ws_behavior.column_dimensions['A'].width = 20
            ws_behavior.column_dimensions['B'].width = 60
        
        # Incidents sheet
        if report.incidents:
            ws_incidents = wb.create_sheet('الحوادث')
            ws_incidents.right_to_left = True
            
            ws_incidents['A1'] = 'النوع'
            ws_incidents['B1'] = 'الوصف'
            ws_incidents['C1'] = 'التاريخ والوقت'
            ws_incidents['D1'] = 'الموقع'
            
            for cell in ['A1', 'B1', 'C1', 'D1']:
                ws_incidents[cell].font = header_font
                ws_incidents[cell].fill = PatternFill(start_color='d32f2f', end_color='d32f2f', fill_type='solid')
                ws_incidents[cell].border = border
            
            row = 2
            for incident in report.incidents:
                ws_incidents[f'A{row}'] = incident.incident_type.value if incident.incident_type else ''
                ws_incidents[f'B{row}'] = incident.description or ''
                ws_incidents[f'C{row}'] = incident.incident_datetime.strftime('%Y-%m-%d %H:%M') if incident.incident_datetime else ''
                ws_incidents[f'D{row}'] = incident.location or ''
                for cell in [f'A{row}', f'B{row}', f'C{row}', f'D{row}']:
                    ws_incidents[cell].border = border
                row += 1
            
            ws_incidents.column_dimensions['A'].width = 20
            ws_incidents.column_dimensions['B'].width = 40
            ws_incidents.column_dimensions['C'].width = 20
            ws_incidents.column_dimensions['D'].width = 20
        
        wb.save(buffer)
        buffer.seek(0)
        return buffer
