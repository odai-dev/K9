"""
Unified Report Service
======================
Centralized API for all report operations in the K9 Operations Management System.
Supports the Generate Once → Preview → Export pattern with caching and approval workflows.
"""
import json
import logging
import os
from io import BytesIO
from datetime import datetime
from typing import Dict, List, Optional, Tuple, Any, Union

from flask import request
from sqlalchemy import desc
from app import db
from k9.models.models import User, Project, VeterinaryVisit, BreedingTrainingActivity, CaretakerDailyLog
from k9.models.models_handler_daily import HandlerReport, ShiftReport
from k9.models.report_models import (
    ReportContext, ReportDefinition, ReportExportHistory, ReportApprovalHistory,
    UnifiedReportType, UnifiedReportStatus, ExportFormat, ReportPriority,
    REPORT_TYPE_NAMES_AR, REPORT_STATUS_NAMES_AR, SOURCE_TABLE_TO_TYPE
)
from k9.services.permission_service import PermissionService
from k9.services.report_export_service import ReportExportService

from reportlab.lib.pagesizes import A4
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Spacer
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from k9.utils.pdf_minimal_elegant import (
    create_minimal_header,
    create_info_section,
    create_text_section,
    create_data_table,
    get_minimal_styles,
    add_page_number,
    MinimalColors,
    create_spacer
)
from k9.utils.utils_pdf_rtl import rtl, register_arabic_fonts, get_arabic_font_name

logger = logging.getLogger(__name__)


class UnifiedReportService:
    """
    Unified service for all report operations.
    Provides a centralized API for generating, previewing, exporting, and approving reports.
    """
    
    REPORT_TYPE_MODELS = {
        UnifiedReportType.HANDLER: HandlerReport,
        UnifiedReportType.SHIFT: ShiftReport,
        UnifiedReportType.TRAINER: BreedingTrainingActivity,
        UnifiedReportType.VET: VeterinaryVisit,
        UnifiedReportType.CARETAKER: CaretakerDailyLog,
    }
    
    SOURCE_TABLE_NAMES = {
        UnifiedReportType.HANDLER: 'handler_report',
        UnifiedReportType.SHIFT: 'shift_report',
        UnifiedReportType.TRAINER: 'breeding_training_activity',
        UnifiedReportType.VET: 'veterinary_visit',
        UnifiedReportType.CARETAKER: 'caretaker_daily_log',
    }
    
    @classmethod
    def generate_report(
        cls,
        report_type: Union[str, UnifiedReportType],
        user_id: str,
        project_id: Optional[str] = None,
        data_filters: Optional[Dict] = None,
        source_report_id: Optional[str] = None
    ) -> Tuple[Optional[ReportContext], str]:
        """
        Generate a report context with cached data.
        Creates or updates a ReportContext and fetches/caches the report data.
        
        Args:
            report_type: The type of report to generate
            user_id: The user requesting the report
            project_id: Optional project scope
            data_filters: Optional filters (date_from, date_to, dog_id, etc.)
            source_report_id: Optional ID of specific source report
            
        Returns:
            Tuple of (ReportContext or None, message)
        """
        if isinstance(report_type, str):
            try:
                report_type = UnifiedReportType(report_type)
            except ValueError:
                return None, f"نوع التقرير غير صالح: {report_type}"
        
        if not PermissionService.has_permission(user_id, 'reports.view', project_id):
            return None, "ليس لديك صلاحية لعرض التقارير"
        
        try:
            context = ReportContext.query.filter_by(
                report_type=report_type,
                project_id=project_id,
                source_report_id=source_report_id,
                created_by_user_id=user_id
            ).first()
            
            if context and context.cache_valid:
                return context, "تم استرجاع التقرير من الذاكرة المؤقتة"
            
            if not context:
                context = ReportContext(
                    report_type=report_type,
                    project_id=project_id,
                    source_report_id=source_report_id,
                    source_report_table=cls.SOURCE_TABLE_NAMES.get(report_type),
                    created_by_user_id=user_id,
                    status=UnifiedReportStatus.DRAFT,
                    priority=ReportPriority.NORMAL
                )
                
                if data_filters:
                    if 'date_from' in data_filters:
                        context.date_from = data_filters['date_from']
                    if 'date_to' in data_filters:
                        context.date_to = data_filters['date_to']
                    if 'report_date' in data_filters:
                        context.report_date = data_filters['report_date']
                    if 'dog_id' in data_filters:
                        context.dog_id = data_filters['dog_id']
                
                context.title_ar = REPORT_TYPE_NAMES_AR.get(report_type, 'تقرير')
                context.title = context.title_ar
                
                db.session.add(context)
                db.session.flush()
            
            cached_data = cls._fetch_report_data(report_type, source_report_id, data_filters)
            
            context.update_cache(cached_data)
            
            db.session.commit()
            return context, "تم إنشاء التقرير بنجاح"
            
        except Exception as e:
            db.session.rollback()
            logger.error(f"Error generating report: {e}")
            return None, f"خطأ في إنشاء التقرير: {str(e)}"
    
    @classmethod
    def get_report_context(cls, context_id: str, user_id: Optional[str] = None) -> Optional[ReportContext]:
        """
        Retrieve a report context with its cached data.
        
        Args:
            context_id: The ID of the report context
            user_id: Optional user ID for permission check
            
        Returns:
            ReportContext or None
        """
        context = ReportContext.query.get(context_id)
        
        if not context:
            return None
        
        if user_id:
            project_id = str(context.project_id) if context.project_id else None
            if not PermissionService.has_permission(user_id, 'reports.view', project_id):
                if str(context.created_by_user_id) != str(user_id):
                    return None
        
        return context
    
    @classmethod
    def preview_report(cls, context_id: str, user_id: str) -> Tuple[Optional[Dict], str]:
        """
        Return HTML preview data for a report.
        Checks view permissions before returning data.
        
        Args:
            context_id: The report context ID
            user_id: The user requesting the preview
            
        Returns:
            Tuple of (preview data dict or None, message)
        """
        context = cls.get_report_context(context_id, user_id)
        
        if not context:
            return None, "التقرير غير موجود أو ليس لديك صلاحية للوصول"
        
        if not context.cache_valid or not context.cached_data:
            cached_data = cls._fetch_report_data(
                context.report_type,
                str(context.source_report_id) if context.source_report_id else None,
                {
                    'date_from': context.date_from,
                    'date_to': context.date_to,
                    'dog_id': str(context.dog_id) if context.dog_id else None
                }
            )
            context.update_cache(cached_data)
            db.session.commit()
        
        preview_data = {
            'context_id': str(context.id),
            'report_type': context.report_type.value,
            'report_type_name': REPORT_TYPE_NAMES_AR.get(context.report_type, 'تقرير'),
            'status': context.status.value,
            'status_name': REPORT_STATUS_NAMES_AR.get(context.status, 'غير معروف'),
            'title': context.title,
            'title_ar': context.title_ar,
            'cached_data': context.cached_data,
            'cached_at': context.cached_at.isoformat() if context.cached_at else None,
            'created_at': context.created_at.isoformat(),
            'can_export': cls.can_export(context_id, user_id)[0],
            'can_submit': context.status == UnifiedReportStatus.DRAFT
        }
        
        return preview_data, "تم تحميل المعاينة بنجاح"
    
    @classmethod
    def export_pdf(cls, context_id: str, user_id: str) -> Tuple[Optional[BytesIO], str]:
        """
        Export report to PDF format.
        Checks export permissions and approval status.
        
        Args:
            context_id: The report context ID
            user_id: The user requesting the export
            
        Returns:
            Tuple of (BytesIO PDF data or None, message)
        """
        can_export, message = cls.can_export(context_id, user_id)
        if not can_export:
            return None, message
        
        context = cls.get_report_context(context_id, user_id)
        if not context:
            return None, "التقرير غير موجود"
        
        try:
            pdf_buffer = cls._generate_pdf(context)
            
            if pdf_buffer:
                cls._log_export(context_id, user_id, ExportFormat.PDF, True)
                return pdf_buffer, "تم تصدير التقرير بصيغة PDF بنجاح"
            else:
                cls._log_export(context_id, user_id, ExportFormat.PDF, False, "فشل إنشاء ملف PDF")
                return None, "فشل في إنشاء ملف PDF"
                
        except Exception as e:
            logger.error(f"Error exporting PDF: {e}")
            cls._log_export(context_id, user_id, ExportFormat.PDF, False, str(e))
            return None, f"خطأ في التصدير: {str(e)}"
    
    @classmethod
    def export_excel(cls, context_id: str, user_id: str) -> Tuple[Optional[BytesIO], str]:
        """
        Export report to Excel format.
        Checks export permissions and approval status.
        
        Args:
            context_id: The report context ID
            user_id: The user requesting the export
            
        Returns:
            Tuple of (BytesIO Excel data or None, message)
        """
        can_export, message = cls.can_export(context_id, user_id)
        if not can_export:
            return None, message
        
        context = cls.get_report_context(context_id, user_id)
        if not context:
            return None, "التقرير غير موجود"
        
        try:
            excel_buffer = cls._generate_excel(context)
            
            if excel_buffer:
                cls._log_export(context_id, user_id, ExportFormat.EXCEL, True)
                return excel_buffer, "تم تصدير التقرير بصيغة Excel بنجاح"
            else:
                cls._log_export(context_id, user_id, ExportFormat.EXCEL, False, "فشل إنشاء ملف Excel")
                return None, "فشل في إنشاء ملف Excel"
                
        except Exception as e:
            logger.error(f"Error exporting Excel: {e}")
            cls._log_export(context_id, user_id, ExportFormat.EXCEL, False, str(e))
            return None, f"خطأ في التصدير: {str(e)}"
    
    @classmethod
    def submit_for_review(cls, context_id: str, user_id: str) -> Tuple[bool, str]:
        """
        Submit a report for PM review.
        Changes status from DRAFT to SUBMITTED.
        
        Args:
            context_id: The report context ID
            user_id: The user submitting the report
            
        Returns:
            Tuple of (success bool, message)
        """
        context = ReportContext.query.get(context_id)
        
        if not context:
            return False, "التقرير غير موجود"
        
        if str(context.created_by_user_id) != str(user_id):
            return False, "لا يمكنك إرسال تقرير لم تقم بإنشائه"
        
        if context.status != UnifiedReportStatus.DRAFT:
            return False, f"التقرير في حالة {REPORT_STATUS_NAMES_AR.get(context.status)} ولا يمكن إرساله"
        
        try:
            from_status = context.status
            context.submit()
            
            cls._log_approval_action(
                context_id=context_id,
                from_status=from_status,
                to_status=UnifiedReportStatus.SUBMITTED,
                action='submit',
                action_by_user_id=user_id
            )
            
            db.session.commit()
            return True, "تم إرسال التقرير للمراجعة بنجاح"
            
        except Exception as e:
            db.session.rollback()
            logger.error(f"Error submitting report: {e}")
            return False, f"خطأ في إرسال التقرير: {str(e)}"
    
    @classmethod
    def pm_approve(
        cls,
        context_id: str,
        pm_user_id: str,
        notes: Optional[str] = None
    ) -> Tuple[bool, str]:
        """
        PM approves a submitted report.
        Changes status from SUBMITTED to PM_APPROVED or FORWARDED_TO_ADMIN.
        
        Args:
            context_id: The report context ID
            pm_user_id: The PM user ID
            notes: Optional approval notes
            
        Returns:
            Tuple of (success bool, message)
        """
        context = ReportContext.query.get(context_id)
        
        if not context:
            return False, "التقرير غير موجود"
        
        project_id = str(context.project_id) if context.project_id else None
        if not PermissionService.has_permission(pm_user_id, 'pm.reports.approve', project_id):
            return False, "ليس لديك صلاحية لاعتماد التقارير"
        
        if context.status != UnifiedReportStatus.SUBMITTED:
            return False, f"التقرير في حالة {REPORT_STATUS_NAMES_AR.get(context.status)} ولا يمكن اعتماده"
        
        if str(context.created_by_user_id) == str(pm_user_id):
            return False, "لا يمكنك اعتماد تقريرك الخاص"
        
        try:
            from_status = context.status
            context.status = UnifiedReportStatus.PM_APPROVED
            context.pm_reviewed_by_user_id = pm_user_id
            context.pm_reviewed_at = datetime.utcnow()
            context.pm_review_notes = notes
            context.pm_review_status = 'approved'
            context.updated_at = datetime.utcnow()
            
            cls._log_approval_action(
                context_id=context_id,
                from_status=from_status,
                to_status=UnifiedReportStatus.PM_APPROVED,
                action='approve',
                action_by_user_id=pm_user_id,
                notes=notes
            )
            
            db.session.commit()
            return True, "تم اعتماد التقرير بنجاح"
            
        except Exception as e:
            db.session.rollback()
            logger.error(f"Error approving report: {e}")
            return False, f"خطأ في اعتماد التقرير: {str(e)}"
    
    @classmethod
    def pm_reject(
        cls,
        context_id: str,
        pm_user_id: str,
        reason: str
    ) -> Tuple[bool, str]:
        """
        PM rejects a submitted report with reason.
        Changes status from SUBMITTED to PM_REJECTED.
        
        Args:
            context_id: The report context ID
            pm_user_id: The PM user ID
            reason: Rejection reason (required)
            
        Returns:
            Tuple of (success bool, message)
        """
        if not reason or not reason.strip():
            return False, "يجب إدخال سبب الرفض"
        
        context = ReportContext.query.get(context_id)
        
        if not context:
            return False, "التقرير غير موجود"
        
        project_id = str(context.project_id) if context.project_id else None
        if not PermissionService.has_permission(pm_user_id, 'pm.reports.approve', project_id):
            return False, "ليس لديك صلاحية لرفض التقارير"
        
        if context.status != UnifiedReportStatus.SUBMITTED:
            return False, f"التقرير في حالة {REPORT_STATUS_NAMES_AR.get(context.status)} ولا يمكن رفضه"
        
        if str(context.created_by_user_id) == str(pm_user_id):
            return False, "لا يمكنك رفض تقريرك الخاص"
        
        try:
            from_status = context.status
            context.status = UnifiedReportStatus.PM_REJECTED
            context.pm_reviewed_by_user_id = pm_user_id
            context.pm_reviewed_at = datetime.utcnow()
            context.pm_review_notes = reason
            context.pm_review_status = 'rejected'
            context.updated_at = datetime.utcnow()
            
            cls._log_approval_action(
                context_id=context_id,
                from_status=from_status,
                to_status=UnifiedReportStatus.PM_REJECTED,
                action='reject',
                action_by_user_id=pm_user_id,
                notes=reason,
                rejection_reason=reason
            )
            
            db.session.commit()
            return True, "تم رفض التقرير"
            
        except Exception as e:
            db.session.rollback()
            logger.error(f"Error rejecting report: {e}")
            return False, f"خطأ في رفض التقرير: {str(e)}"
    
    @classmethod
    def admin_approve(
        cls,
        context_id: str,
        admin_user_id: str,
        notes: Optional[str] = None
    ) -> Tuple[bool, str]:
        """
        Admin gives final approval to a report.
        Changes status to APPROVED.
        
        Args:
            context_id: The report context ID
            admin_user_id: The admin user ID
            notes: Optional approval notes
            
        Returns:
            Tuple of (success bool, message)
        """
        context = ReportContext.query.get(context_id)
        
        if not context:
            return False, "التقرير غير موجود"
        
        if not PermissionService.has_any_permission(admin_user_id, ['admin.*', 'admin.reports.approve']):
            return False, "ليس لديك صلاحية الإدارة العامة"
        
        allowed_statuses = [
            UnifiedReportStatus.PM_APPROVED,
            UnifiedReportStatus.FORWARDED_TO_ADMIN,
            UnifiedReportStatus.SUBMITTED
        ]
        if context.status not in allowed_statuses:
            return False, f"التقرير في حالة {REPORT_STATUS_NAMES_AR.get(context.status)} ولا يمكن اعتماده"
        
        try:
            from_status = context.status
            context.status = UnifiedReportStatus.APPROVED
            context.admin_reviewed_by_user_id = admin_user_id
            context.admin_reviewed_at = datetime.utcnow()
            context.admin_review_notes = notes
            context.admin_review_status = 'approved'
            context.updated_at = datetime.utcnow()
            
            cls._log_approval_action(
                context_id=context_id,
                from_status=from_status,
                to_status=UnifiedReportStatus.APPROVED,
                action='admin_approve',
                action_by_user_id=admin_user_id,
                notes=notes
            )
            
            db.session.commit()
            return True, "تم الاعتماد النهائي للتقرير"
            
        except Exception as e:
            db.session.rollback()
            logger.error(f"Error admin approving report: {e}")
            return False, f"خطأ في الاعتماد: {str(e)}"
    
    @classmethod
    def can_export(cls, context_id: str, user_id: str) -> Tuple[bool, str]:
        """
        Check if a user can export a report.
        - Admin: Can always export
        - PM: Needs report to be approved
        - Creator: Needs report to be approved or in draft
        
        Args:
            context_id: The report context ID
            user_id: The user ID
            
        Returns:
            Tuple of (can_export bool, message)
        """
        context = ReportContext.query.get(context_id)
        
        if not context:
            return False, "التقرير غير موجود"
        
        project_id = str(context.project_id) if context.project_id else None
        
        if PermissionService.has_any_permission(user_id, ['admin.*', 'admin.reports.export']):
            return True, "يمكن للمشرف تصدير جميع التقارير"
        
        if not PermissionService.has_permission(user_id, 'reports.export', project_id):
            return False, "ليس لديك صلاحية تصدير التقارير"
        
        is_creator = str(context.created_by_user_id) == str(user_id)
        
        approved_statuses = [
            UnifiedReportStatus.APPROVED,
            UnifiedReportStatus.PM_APPROVED,
            UnifiedReportStatus.FORWARDED_TO_ADMIN
        ]
        
        if context.status in approved_statuses:
            return True, "التقرير معتمد ويمكن تصديره"
        
        if is_creator and context.status == UnifiedReportStatus.DRAFT:
            return True, "يمكنك تصدير مسودة تقريرك"
        
        if is_creator:
            return True, "يمكنك تصدير تقريرك"
        
        return False, "التقرير غير معتمد ولا يمكن تصديره"
    
    @classmethod
    def get_pending_reports_for_pm(
        cls,
        pm_user_id: str,
        project_id: Optional[str] = None,
        limit: int = 50
    ) -> List[ReportContext]:
        """
        Get all reports pending PM review for a specific PM or project.
        
        Args:
            pm_user_id: The PM user ID requesting the reports
            project_id: Optional project ID to filter by
            limit: Maximum number of reports to return
            
        Returns:
            List of ReportContext objects pending PM review
        """
        query = ReportContext.query.filter(
            ReportContext.status == UnifiedReportStatus.SUBMITTED
        )
        
        if project_id:
            query = query.filter(ReportContext.project_id == project_id)
        
        query = query.filter(ReportContext.created_by_user_id != pm_user_id)
        
        query = query.order_by(desc(ReportContext.submitted_at))
        
        return query.limit(limit).all()
    
    @classmethod
    def get_pending_reports_count_for_pm(
        cls,
        pm_user_id: str,
        project_id: Optional[str] = None
    ) -> int:
        """
        Get count of reports pending PM review.
        
        Args:
            pm_user_id: The PM user ID
            project_id: Optional project ID to filter by
            
        Returns:
            Count of pending reports
        """
        query = ReportContext.query.filter(
            ReportContext.status == UnifiedReportStatus.SUBMITTED
        )
        
        if project_id:
            query = query.filter(ReportContext.project_id == project_id)
        
        query = query.filter(ReportContext.created_by_user_id != pm_user_id)
        
        return query.count()
    
    @classmethod
    def invalidate_cache(cls, context_id: str) -> bool:
        """
        Invalidate cached data for a report context.
        Call this when source data changes.
        
        Args:
            context_id: The report context ID
            
        Returns:
            Success bool
        """
        context = ReportContext.query.get(context_id)
        if context:
            context.invalidate_cache()
            db.session.commit()
            return True
        return False
    
    @classmethod
    def invalidate_cache_by_source(cls, source_table: str, source_id: str) -> int:
        """
        Invalidate all cached contexts linked to a source report.
        
        Args:
            source_table: The source table name
            source_id: The source report ID
            
        Returns:
            Number of contexts invalidated
        """
        contexts = ReportContext.query.filter_by(
            source_report_table=source_table,
            source_report_id=source_id
        ).all()
        
        count = 0
        for context in contexts:
            context.invalidate_cache()
            count += 1
        
        if count > 0:
            db.session.commit()
        
        return count
    
    @classmethod
    def _fetch_report_data(
        cls,
        report_type: UnifiedReportType,
        source_report_id: Optional[str] = None,
        data_filters: Optional[Dict] = None
    ) -> Dict[str, Any]:
        """
        Fetch report data based on type.
        Routes to appropriate type-specific fetch method.
        """
        if report_type == UnifiedReportType.HANDLER:
            return cls._fetch_handler_report_data(source_report_id, data_filters)
        elif report_type == UnifiedReportType.SHIFT:
            return cls._fetch_shift_report_data(source_report_id, data_filters)
        elif report_type == UnifiedReportType.VET:
            return cls._fetch_vet_report_data(source_report_id, data_filters)
        elif report_type == UnifiedReportType.TRAINER:
            return cls._fetch_trainer_report_data(source_report_id, data_filters)
        elif report_type == UnifiedReportType.CARETAKER:
            return cls._fetch_caretaker_report_data(source_report_id, data_filters)
        else:
            return {'error': f'Unsupported report type: {report_type.value}'}
    
    @classmethod
    def _fetch_handler_report_data(cls, report_id: Optional[str], filters: Optional[Dict] = None) -> Dict[str, Any]:
        """Fetch handler daily report data"""
        if report_id:
            report = HandlerReport.query.get(report_id)
            if not report:
                return {'error': 'Handler report not found'}
            
            return {
                'id': str(report.id),
                'date': report.date.isoformat() if report.date else None,
                'report_type': report.report_type.value if report.report_type else None,
                'status': report.status if isinstance(report.status, str) else report.status.value if report.status else None,
                'handler': {
                    'id': str(report.handler_user_id),
                    'name': report.handler.username if report.handler else None
                } if report.handler_user_id else None,
                'dog': {
                    'id': str(report.dog_id),
                    'name': report.dog.name if report.dog else None,
                    'code': report.dog.code if report.dog else None
                } if report.dog_id else None,
                'project': {
                    'id': str(report.project_id),
                    'name': report.project.name if report.project else None
                } if report.project_id else None,
                'location': report.location,
                'health': cls._serialize_health(report.health) if hasattr(report, 'health') and report.health else None,
                'care': cls._serialize_care(report.care) if hasattr(report, 'care') and report.care else None,
                'behavior': cls._serialize_behavior(report.behavior) if hasattr(report, 'behavior') and report.behavior else None,
                'training_sessions': [cls._serialize_training(s) for s in report.training_sessions] if hasattr(report, 'training_sessions') else [],
                'incidents': [cls._serialize_incident(i) for i in report.incidents] if hasattr(report, 'incidents') else [],
                'notes': report.notes if hasattr(report, 'notes') else None,
                'submitted_at': report.submitted_at.isoformat() if report.submitted_at else None,
                'reviewed_at': report.reviewed_at.isoformat() if hasattr(report, 'reviewed_at') and report.reviewed_at else None,
                'created_at': report.created_at.isoformat() if report.created_at else None
            }
        
        return {'reports': [], 'total': 0}
    
    @classmethod
    def _fetch_shift_report_data(cls, report_id: Optional[str], filters: Optional[Dict] = None) -> Dict[str, Any]:
        """Fetch shift report data"""
        if report_id:
            report = ShiftReport.query.get(report_id)
            if not report:
                return {'error': 'Shift report not found'}
            
            return {
                'id': str(report.id),
                'date': report.date.isoformat() if report.date else None,
                'shift_id': str(report.shift_id) if report.shift_id else None,
                'status': report.status if isinstance(report.status, str) else report.status.value if report.status else None,
                'handler': {
                    'id': str(report.handler_user_id),
                    'name': report.handler.username if report.handler else None
                } if report.handler_user_id else None,
                'dog': {
                    'id': str(report.dog_id),
                    'name': report.dog.name if report.dog else None,
                    'code': report.dog.code if report.dog else None
                } if report.dog_id else None,
                'project': {
                    'id': str(report.project_id),
                    'name': report.project.name if report.project else None
                } if report.project_id else None,
                'location': report.location if hasattr(report, 'location') else None,
                'quick_health_check': report.quick_health_check if hasattr(report, 'quick_health_check') else None,
                'quick_health_notes': report.quick_health_notes if hasattr(report, 'quick_health_notes') else None,
                'incidents': [cls._serialize_incident(i) for i in report.incidents] if hasattr(report, 'incidents') and report.incidents else [],
                'notes': report.notes if hasattr(report, 'notes') else None,
                'submitted_at': report.submitted_at.isoformat() if report.submitted_at else None,
                'created_at': report.created_at.isoformat() if report.created_at else None
            }
        
        return {'reports': [], 'total': 0}
    
    @classmethod
    def _fetch_vet_report_data(cls, report_id: Optional[str], filters: Optional[Dict] = None) -> Dict[str, Any]:
        """Fetch veterinary visit report data"""
        if report_id:
            report = VeterinaryVisit.query.get(report_id)
            if not report:
                return {'error': 'Veterinary report not found'}
            
            return {
                'id': str(report.id),
                'visit_date': report.visit_date.isoformat() if hasattr(report, 'visit_date') and report.visit_date else None,
                'visit_type': report.visit_type.value if hasattr(report, 'visit_type') and report.visit_type else None,
                'status': report.status if isinstance(report.status, str) else getattr(report.status, 'value', None) if report.status else None,
                'dog': {
                    'id': str(report.dog_id),
                    'name': report.dog.name if report.dog else None,
                    'code': report.dog.code if report.dog else None
                } if report.dog_id else None,
                'project': {
                    'id': str(report.project_id),
                    'name': report.project.name if report.project else None
                } if report.project_id else None,
                'diagnosis': report.diagnosis if hasattr(report, 'diagnosis') else None,
                'treatment': report.treatment if hasattr(report, 'treatment') else None,
                'medications': report.medications if hasattr(report, 'medications') else None,
                'notes': report.notes if hasattr(report, 'notes') else None,
                'follow_up_date': report.follow_up_date.isoformat() if hasattr(report, 'follow_up_date') and report.follow_up_date else None,
                'created_by': {
                    'id': str(report.created_by_user_id),
                    'name': report.created_by.username if hasattr(report, 'created_by') and report.created_by else None
                } if hasattr(report, 'created_by_user_id') and report.created_by_user_id else None,
                'submitted_at': report.submitted_at.isoformat() if hasattr(report, 'submitted_at') and report.submitted_at else None,
                'created_at': report.created_at.isoformat() if report.created_at else None
            }
        
        return {'reports': [], 'total': 0}
    
    @classmethod
    def _fetch_trainer_report_data(cls, report_id: Optional[str], filters: Optional[Dict] = None) -> Dict[str, Any]:
        """Fetch trainer (breeding training activity) report data"""
        if report_id:
            report = BreedingTrainingActivity.query.get(report_id)
            if not report:
                return {'error': 'Trainer report not found'}
            
            return {
                'id': str(report.id),
                'activity_date': report.activity_date.isoformat() if hasattr(report, 'activity_date') and report.activity_date else None,
                'category': report.category.value if hasattr(report, 'category') and report.category else None,
                'status': report.status if isinstance(report.status, str) else getattr(report.status, 'value', None) if report.status else None,
                'dog': {
                    'id': str(report.dog_id),
                    'name': report.dog.name if report.dog else None,
                    'code': report.dog.code if report.dog else None
                } if report.dog_id else None,
                'project': {
                    'id': str(report.project_id),
                    'name': report.project.name if report.project else None
                } if report.project_id else None,
                'description': report.description if hasattr(report, 'description') else None,
                'duration_minutes': report.duration_minutes if hasattr(report, 'duration_minutes') else None,
                'performance_rating': report.performance_rating.value if hasattr(report, 'performance_rating') and report.performance_rating else None,
                'notes': report.notes if hasattr(report, 'notes') else None,
                'created_by': {
                    'id': str(report.created_by_user_id),
                    'name': report.created_by.username if hasattr(report, 'created_by') and report.created_by else None
                } if hasattr(report, 'created_by_user_id') and report.created_by_user_id else None,
                'submitted_at': report.submitted_at.isoformat() if hasattr(report, 'submitted_at') and report.submitted_at else None,
                'created_at': report.created_at.isoformat() if report.created_at else None
            }
        
        return {'reports': [], 'total': 0}
    
    @classmethod
    def _fetch_caretaker_report_data(cls, report_id: Optional[str], filters: Optional[Dict] = None) -> Dict[str, Any]:
        """Fetch caretaker daily log report data"""
        if report_id:
            report = CaretakerDailyLog.query.get(report_id)
            if not report:
                return {'error': 'Caretaker report not found'}
            
            return {
                'id': str(report.id),
                'log_date': report.log_date.isoformat() if hasattr(report, 'log_date') and report.log_date else None,
                'status': report.status if isinstance(report.status, str) else getattr(report.status, 'value', None) if report.status else None,
                'dog': {
                    'id': str(report.dog_id),
                    'name': report.dog.name if report.dog else None,
                    'code': report.dog.code if report.dog else None
                } if report.dog_id else None,
                'project': {
                    'id': str(report.project_id),
                    'name': report.project.name if report.project else None
                } if report.project_id else None,
                'feeding_done': report.feeding_done if hasattr(report, 'feeding_done') else None,
                'feeding_notes': report.feeding_notes if hasattr(report, 'feeding_notes') else None,
                'cleaning_done': report.cleaning_done if hasattr(report, 'cleaning_done') else None,
                'cleaning_notes': report.cleaning_notes if hasattr(report, 'cleaning_notes') else None,
                'health_observation': report.health_observation if hasattr(report, 'health_observation') else None,
                'notes': report.notes if hasattr(report, 'notes') else None,
                'created_by': {
                    'id': str(report.created_by_user_id),
                    'name': report.created_by.username if hasattr(report, 'created_by') and report.created_by else None
                } if hasattr(report, 'created_by_user_id') and report.created_by_user_id else None,
                'submitted_at': report.submitted_at.isoformat() if hasattr(report, 'submitted_at') and report.submitted_at else None,
                'created_at': report.created_at.isoformat() if report.created_at else None
            }
        
        return {'reports': [], 'total': 0}
    
    @classmethod
    def _serialize_health(cls, health) -> Optional[Dict]:
        """Serialize health check data"""
        if not health:
            return None
        return {
            'eyes_status': health.eyes_status.value if hasattr(health, 'eyes_status') and health.eyes_status else None,
            'eyes_notes': health.eyes_notes if hasattr(health, 'eyes_notes') else None,
            'nose_status': health.nose_status.value if hasattr(health, 'nose_status') and health.nose_status else None,
            'nose_notes': health.nose_notes if hasattr(health, 'nose_notes') else None,
            'ears_status': health.ears_status.value if hasattr(health, 'ears_status') and health.ears_status else None,
            'ears_notes': health.ears_notes if hasattr(health, 'ears_notes') else None,
            'mouth_status': health.mouth_status.value if hasattr(health, 'mouth_status') and health.mouth_status else None,
            'mouth_notes': health.mouth_notes if hasattr(health, 'mouth_notes') else None,
        }
    
    @classmethod
    def _serialize_care(cls, care) -> Optional[Dict]:
        """Serialize care data"""
        if not care:
            return None
        return {
            'food_amount': care.food_amount if hasattr(care, 'food_amount') else None,
            'food_type': care.food_type if hasattr(care, 'food_type') else None,
            'water_amount': care.water_amount if hasattr(care, 'water_amount') else None,
            'grooming_done': care.grooming_done if hasattr(care, 'grooming_done') else None,
            'washing_done': care.washing_done if hasattr(care, 'washing_done') else None,
        }
    
    @classmethod
    def _serialize_behavior(cls, behavior) -> Optional[Dict]:
        """Serialize behavior data"""
        if not behavior:
            return None
        return {
            'good_behavior_notes': behavior.good_behavior_notes if hasattr(behavior, 'good_behavior_notes') else None,
            'bad_behavior_notes': behavior.bad_behavior_notes if hasattr(behavior, 'bad_behavior_notes') else None,
        }
    
    @classmethod
    def _serialize_training(cls, session) -> Dict:
        """Serialize training session data"""
        return {
            'training_type': session.training_type.value if hasattr(session, 'training_type') and session.training_type else None,
            'description': session.description if hasattr(session, 'description') else None,
            'time_from': session.time_from.isoformat() if hasattr(session, 'time_from') and session.time_from else None,
            'time_to': session.time_to.isoformat() if hasattr(session, 'time_to') and session.time_to else None,
            'notes': session.notes if hasattr(session, 'notes') else None,
        }
    
    @classmethod
    def _serialize_incident(cls, incident) -> Dict:
        """Serialize incident data"""
        return {
            'incident_type': incident.incident_type.value if hasattr(incident, 'incident_type') and incident.incident_type else None,
            'description': incident.description if hasattr(incident, 'description') else None,
            'incident_datetime': incident.incident_datetime.isoformat() if hasattr(incident, 'incident_datetime') and incident.incident_datetime else None,
            'location': incident.location if hasattr(incident, 'location') else None,
        }
    
    EXPORT_METHOD_MAP_PDF = {
        UnifiedReportType.HANDLER: 'export_handler_report_to_pdf',
        UnifiedReportType.SHIFT: 'export_shift_report_to_pdf',
        UnifiedReportType.VET: 'export_vet_report_to_pdf',
        UnifiedReportType.TRAINER: 'export_trainer_report_to_pdf',
        UnifiedReportType.CARETAKER: 'export_caretaker_report_to_pdf',
    }
    
    EXPORT_METHOD_MAP_EXCEL = {
        UnifiedReportType.HANDLER: 'export_handler_report_to_excel',
        UnifiedReportType.SHIFT: 'export_shift_report_to_excel',
        UnifiedReportType.VET: 'export_vet_report_to_excel',
        UnifiedReportType.TRAINER: 'export_trainer_report_to_excel',
        UnifiedReportType.CARETAKER: 'export_caretaker_report_to_excel',
    }
    
    @classmethod
    def _generate_pdf(cls, context: ReportContext) -> Optional[BytesIO]:
        """
        Generate PDF from report context cached data using the minimal elegant design.
        Creates a unified professional PDF layout for all report types.
        """
        if not context.cached_data and not context.source_report_id:
            return None
        
        if context.source_report_id:
            method_name = cls.EXPORT_METHOD_MAP_PDF.get(context.report_type)
            if method_name and hasattr(ReportExportService, method_name):
                export_method = getattr(ReportExportService, method_name)
                result = export_method(str(context.source_report_id))
                if result:
                    return result
        
        try:
            register_arabic_fonts()
            buffer = BytesIO()
            
            doc = SimpleDocTemplate(
                buffer,
                pagesize=A4,
                rightMargin=2*cm,
                leftMargin=2*cm,
                topMargin=2*cm,
                bottomMargin=2.5*cm
            )
            
            styles = get_minimal_styles()
            story = []
            
            logo_path = 'k9/static/img/company_logo.png'
            if not os.path.exists(logo_path):
                logo_path = None
            
            report_title = context.title_ar or REPORT_TYPE_NAMES_AR.get(context.report_type, 'تقرير')
            
            metadata = {}
            if context.report_date:
                metadata['date'] = context.report_date.strftime('%Y-%m-%d')
            elif context.date_from and context.date_to:
                metadata['period'] = f"{context.date_from.strftime('%Y-%m-%d')} - {context.date_to.strftime('%Y-%m-%d')}"
            
            if context.project:
                metadata['project'] = context.project.name
            
            cached_data = context.cached_data or {}
            
            if cached_data.get('handler'):
                handler_info = cached_data.get('handler', {})
                if handler_info.get('name'):
                    metadata['handler'] = handler_info.get('name')
            
            header_elements = create_minimal_header(report_title, metadata, logo_path)
            story.extend(header_elements)
            
            report_meta = {}
            
            status_display = REPORT_STATUS_NAMES_AR.get(context.status, context.status.value if context.status else 'غير معروف')
            report_meta['الحالة'] = status_display
            
            if context.created_by:
                report_meta['مقدم التقرير'] = context.created_by.username
            
            if context.submitted_at:
                report_meta['تاريخ الإرسال'] = context.submitted_at.strftime('%Y-%m-%d %H:%M')
            
            if context.pm_reviewer:
                report_meta['تمت المراجعة بواسطة'] = context.pm_reviewer.username
                if context.pm_reviewed_at:
                    report_meta['تاريخ المراجعة'] = context.pm_reviewed_at.strftime('%Y-%m-%d %H:%M')
            
            if context.pm_review_notes:
                report_meta['ملاحظات المراجعة'] = context.pm_review_notes
            
            if report_meta:
                meta_elements = create_info_section('معلومات التقرير', report_meta)
                story.extend(meta_elements)
            
            cls._add_report_data_to_pdf(story, context.report_type, cached_data, styles)
            
            if context.notes:
                notes_elements = create_text_section('ملاحظات إضافية', context.notes)
                story.extend(notes_elements)
            
            doc.build(story, onFirstPage=add_page_number, onLaterPages=add_page_number)
            buffer.seek(0)
            return buffer
            
        except Exception as e:
            logger.error(f"Error generating unified PDF: {e}")
            return None
    
    @classmethod
    def _add_report_data_to_pdf(cls, story: list, report_type: UnifiedReportType, data: Dict, styles: Dict):
        """Add report-type specific data sections to PDF story"""
        if not data or data.get('error'):
            return
        
        if report_type == UnifiedReportType.HANDLER:
            cls._add_handler_data_to_pdf(story, data, styles)
        elif report_type == UnifiedReportType.SHIFT:
            cls._add_shift_data_to_pdf(story, data, styles)
        elif report_type == UnifiedReportType.VET:
            cls._add_vet_data_to_pdf(story, data, styles)
        elif report_type == UnifiedReportType.TRAINER:
            cls._add_trainer_data_to_pdf(story, data, styles)
        elif report_type == UnifiedReportType.CARETAKER:
            cls._add_caretaker_data_to_pdf(story, data, styles)
        else:
            cls._add_generic_data_to_pdf(story, data, styles)
    
    @classmethod
    def _add_handler_data_to_pdf(cls, story: list, data: Dict, styles: Dict):
        """Add handler report specific data to PDF"""
        dog_info = data.get('dog', {})
        if dog_info:
            dog_section = {
                'اسم الكلب': dog_info.get('name', ''),
                'رمز الكلب': dog_info.get('code', ''),
            }
            story.extend(create_info_section('معلومات الكلب', dog_section))
        
        if data.get('location'):
            story.extend(create_info_section('الموقع', {'الموقع': data.get('location')}))
        
        health = data.get('health')
        if health:
            health_section = {}
            health_fields = [
                ('eyes_status', 'eyes_notes', 'العيون'),
                ('nose_status', 'nose_notes', 'الأنف'),
                ('ears_status', 'ears_notes', 'الأذنين'),
                ('mouth_status', 'mouth_notes', 'الفم'),
            ]
            for status_field, notes_field, label in health_fields:
                status = health.get(status_field)
                notes = health.get(notes_field)
                if status:
                    value = cls._get_health_status_ar(status)
                    if notes:
                        value += f" - {notes}"
                    health_section[label] = value
            
            if health_section:
                story.extend(create_info_section('الفحص الصحي', health_section))
        
        care = data.get('care')
        if care:
            care_section = {}
            if care.get('food_amount'):
                care_section['كمية الطعام'] = care.get('food_amount')
            if care.get('food_type'):
                care_section['نوع الطعام'] = care.get('food_type')
            if care.get('water_amount'):
                care_section['كمية الماء'] = care.get('water_amount')
            if care.get('grooming_done') is not None:
                care_section['التمشيط'] = 'نعم' if care.get('grooming_done') else 'لا'
            if care.get('washing_done') is not None:
                care_section['الغسل'] = 'نعم' if care.get('washing_done') else 'لا'
            
            if care_section:
                story.extend(create_info_section('العناية', care_section))
        
        behavior = data.get('behavior')
        if behavior:
            behavior_section = {}
            if behavior.get('good_behavior_notes'):
                behavior_section['السلوك الإيجابي'] = behavior.get('good_behavior_notes')
            if behavior.get('bad_behavior_notes'):
                behavior_section['السلوك السلبي'] = behavior.get('bad_behavior_notes')
            
            if behavior_section:
                story.extend(create_info_section('السلوك', behavior_section))
        
        training = data.get('training_sessions', [])
        if training:
            headers = ['نوع التدريب', 'الوصف', 'من', 'إلى', 'ملاحظات']
            rows = []
            for session in training:
                rows.append([
                    session.get('training_type', ''),
                    session.get('description', ''),
                    session.get('time_from', '') or '',
                    session.get('time_to', '') or '',
                    session.get('notes', '') or ''
                ])
            if rows:
                story.append(create_spacer(0.5))
                from reportlab.platypus import Paragraph
                story.append(Paragraph(rtl('جلسات التدريب'), styles['SectionHeading']))
                table = create_data_table(headers, rows, [3*cm, 4*cm, 2.5*cm, 2.5*cm, 4*cm])
                story.append(table)
        
        incidents = data.get('incidents', [])
        if incidents:
            headers = ['النوع', 'الوصف', 'التاريخ/الوقت', 'الموقع']
            rows = []
            for incident in incidents:
                rows.append([
                    incident.get('incident_type', ''),
                    incident.get('description', ''),
                    incident.get('incident_datetime', '') or '',
                    incident.get('location', '') or ''
                ])
            if rows:
                story.append(create_spacer(0.5))
                from reportlab.platypus import Paragraph
                story.append(Paragraph(rtl('الحوادث والاشتباهات'), styles['SectionHeading']))
                table = create_data_table(headers, rows, [3*cm, 6*cm, 4*cm, 3*cm])
                story.append(table)
    
    @classmethod
    def _add_shift_data_to_pdf(cls, story: list, data: Dict, styles: Dict):
        """Add shift report specific data to PDF"""
        dog_info = data.get('dog', {})
        if dog_info:
            dog_section = {
                'اسم الكلب': dog_info.get('name', ''),
                'رمز الكلب': dog_info.get('code', ''),
            }
            story.extend(create_info_section('معلومات الكلب', dog_section))
        
        shift_info = {}
        if data.get('location'):
            shift_info['الموقع'] = data.get('location')
        if data.get('quick_health_check'):
            shift_info['الفحص الصحي السريع'] = data.get('quick_health_check')
        if data.get('quick_health_notes'):
            shift_info['ملاحظات صحية'] = data.get('quick_health_notes')
        
        if shift_info:
            story.extend(create_info_section('معلومات الوردية', shift_info))
        
        incidents = data.get('incidents', [])
        if incidents:
            headers = ['النوع', 'الوصف', 'التاريخ/الوقت', 'الموقع']
            rows = []
            for incident in incidents:
                rows.append([
                    incident.get('incident_type', ''),
                    incident.get('description', ''),
                    incident.get('incident_datetime', '') or '',
                    incident.get('location', '') or ''
                ])
            if rows:
                story.append(create_spacer(0.5))
                from reportlab.platypus import Paragraph
                story.append(Paragraph(rtl('الحوادث'), styles['SectionHeading']))
                table = create_data_table(headers, rows)
                story.append(table)
        
        if data.get('notes'):
            story.extend(create_text_section('ملاحظات', data.get('notes')))
    
    @classmethod
    def _add_vet_data_to_pdf(cls, story: list, data: Dict, styles: Dict):
        """Add veterinary report specific data to PDF"""
        dog_info = data.get('dog', {})
        if dog_info:
            dog_section = {
                'اسم الكلب': dog_info.get('name', ''),
                'رمز الكلب': dog_info.get('code', ''),
            }
            story.extend(create_info_section('معلومات الكلب', dog_section))
        
        vet_info = {}
        if data.get('visit_date'):
            vet_info['تاريخ الزيارة'] = data.get('visit_date')
        if data.get('visit_type'):
            vet_info['نوع الزيارة'] = data.get('visit_type')
        if data.get('diagnosis'):
            vet_info['التشخيص'] = data.get('diagnosis')
        if data.get('treatment'):
            vet_info['العلاج'] = data.get('treatment')
        if data.get('medications'):
            vet_info['الأدوية'] = data.get('medications')
        if data.get('follow_up_date'):
            vet_info['موعد المتابعة'] = data.get('follow_up_date')
        
        if vet_info:
            story.extend(create_info_section('تفاصيل الزيارة البيطرية', vet_info))
        
        if data.get('notes'):
            story.extend(create_text_section('ملاحظات', data.get('notes')))
    
    @classmethod
    def _add_trainer_data_to_pdf(cls, story: list, data: Dict, styles: Dict):
        """Add trainer report specific data to PDF"""
        dog_info = data.get('dog', {})
        if dog_info:
            dog_section = {
                'اسم الكلب': dog_info.get('name', ''),
                'رمز الكلب': dog_info.get('code', ''),
            }
            story.extend(create_info_section('معلومات الكلب', dog_section))
        
        training_info = {}
        if data.get('activity_date'):
            training_info['تاريخ النشاط'] = data.get('activity_date')
        if data.get('category'):
            training_info['الفئة'] = data.get('category')
        if data.get('description'):
            training_info['الوصف'] = data.get('description')
        if data.get('duration_minutes'):
            training_info['المدة (دقائق)'] = str(data.get('duration_minutes'))
        if data.get('performance_rating'):
            training_info['تقييم الأداء'] = data.get('performance_rating')
        
        if training_info:
            story.extend(create_info_section('تفاصيل التدريب', training_info))
        
        if data.get('notes'):
            story.extend(create_text_section('ملاحظات', data.get('notes')))
    
    @classmethod
    def _add_caretaker_data_to_pdf(cls, story: list, data: Dict, styles: Dict):
        """Add caretaker report specific data to PDF"""
        dog_info = data.get('dog', {})
        if dog_info:
            dog_section = {
                'اسم الكلب': dog_info.get('name', ''),
                'رمز الكلب': dog_info.get('code', ''),
            }
            story.extend(create_info_section('معلومات الكلب', dog_section))
        
        care_info = {}
        if data.get('log_date'):
            care_info['تاريخ السجل'] = data.get('log_date')
        if data.get('feeding_done') is not None:
            care_info['تم التغذية'] = 'نعم' if data.get('feeding_done') else 'لا'
        if data.get('feeding_notes'):
            care_info['ملاحظات التغذية'] = data.get('feeding_notes')
        if data.get('cleaning_done') is not None:
            care_info['تم التنظيف'] = 'نعم' if data.get('cleaning_done') else 'لا'
        if data.get('cleaning_notes'):
            care_info['ملاحظات التنظيف'] = data.get('cleaning_notes')
        if data.get('health_observation'):
            care_info['ملاحظات صحية'] = data.get('health_observation')
        
        if care_info:
            story.extend(create_info_section('تفاصيل الرعاية اليومية', care_info))
        
        if data.get('notes'):
            story.extend(create_text_section('ملاحظات', data.get('notes')))
    
    @classmethod
    def _add_generic_data_to_pdf(cls, story: list, data: Dict, styles: Dict):
        """Add generic data section for unknown report types"""
        display_data = {}
        skip_keys = ['id', 'error', 'reports', 'total']
        
        for key, value in data.items():
            if key in skip_keys:
                continue
            if value is not None and value != '':
                if isinstance(value, dict):
                    continue
                elif isinstance(value, list):
                    continue
                else:
                    display_data[key] = str(value)
        
        if display_data:
            story.extend(create_info_section('بيانات التقرير', display_data))
    
    @classmethod
    def _get_health_status_ar(cls, status: str) -> str:
        """Get Arabic health status text"""
        status_map = {
            'NORMAL': 'طبيعي',
            'ABNORMAL': 'غير طبيعي',
            'NOT_CHECKED': 'لم يتم الفحص',
            'GOOD': 'جيد',
            'BAD': 'سيء',
        }
        return status_map.get(status, status or '')
    
    @classmethod
    def _generate_excel(cls, context: ReportContext) -> Optional[BytesIO]:
        """
        Generate Excel from report context cached data with professional formatting.
        Creates a unified professional Excel layout for all report types.
        """
        if not context.cached_data and not context.source_report_id:
            return None
        
        if context.source_report_id:
            method_name = cls.EXPORT_METHOD_MAP_EXCEL.get(context.report_type)
            if method_name and hasattr(ReportExportService, method_name):
                export_method = getattr(ReportExportService, method_name)
                result = export_method(str(context.source_report_id))
                if result:
                    return result
        
        try:
            buffer = BytesIO()
            wb = Workbook()
            
            header_font = Font(bold=True, size=12, color='FFFFFF')
            header_fill = PatternFill(start_color='3A6EA5', end_color='3A6EA5', fill_type='solid')
            title_font = Font(bold=True, size=14, color='333333')
            meta_fill = PatternFill(start_color='F5F5F5', end_color='F5F5F5', fill_type='solid')
            border = Border(
                left=Side(style='thin', color='DDDDDD'),
                right=Side(style='thin', color='DDDDDD'),
                top=Side(style='thin', color='DDDDDD'),
                bottom=Side(style='thin', color='DDDDDD')
            )
            
            report_type_name = REPORT_TYPE_NAMES_AR.get(context.report_type, 'تقرير')
            sheet_name = report_type_name[:31]
            
            ws = wb.active
            ws.title = sheet_name
            ws.right_to_left = True
            
            row = 1
            ws.merge_cells(f'A{row}:D{row}')
            ws[f'A{row}'] = context.title_ar or report_type_name
            ws[f'A{row}'].font = title_font
            ws[f'A{row}'].alignment = Alignment(horizontal='center', vertical='center')
            row += 2
            
            meta_data = []
            
            if context.report_date:
                meta_data.append(('التاريخ', context.report_date.strftime('%Y-%m-%d')))
            elif context.date_from and context.date_to:
                meta_data.append(('الفترة', f"{context.date_from.strftime('%Y-%m-%d')} - {context.date_to.strftime('%Y-%m-%d')}"))
            
            if context.project:
                meta_data.append(('المشروع', context.project.name))
            
            status_display = REPORT_STATUS_NAMES_AR.get(context.status, context.status.value if context.status else 'غير معروف')
            meta_data.append(('الحالة', status_display))
            
            if context.created_by:
                meta_data.append(('مقدم التقرير', context.created_by.username))
            
            if context.submitted_at:
                meta_data.append(('تاريخ الإرسال', context.submitted_at.strftime('%Y-%m-%d %H:%M')))
            
            if context.pm_reviewer:
                meta_data.append(('المراجع', context.pm_reviewer.username))
                if context.pm_reviewed_at:
                    meta_data.append(('تاريخ المراجعة', context.pm_reviewed_at.strftime('%Y-%m-%d %H:%M')))
            
            for label, value in meta_data:
                ws[f'A{row}'] = label
                ws[f'B{row}'] = value
                ws[f'A{row}'].font = Font(bold=True)
                ws[f'A{row}'].fill = meta_fill
                ws[f'A{row}'].border = border
                ws[f'B{row}'].border = border
                ws[f'A{row}'].alignment = Alignment(horizontal='right')
                ws[f'B{row}'].alignment = Alignment(horizontal='right')
                row += 1
            
            row += 1
            
            cached_data = context.cached_data or {}
            row = cls._add_report_data_to_excel(ws, row, context.report_type, cached_data, header_font, header_fill, border)
            
            ws.column_dimensions['A'].width = 25
            ws.column_dimensions['B'].width = 35
            ws.column_dimensions['C'].width = 25
            ws.column_dimensions['D'].width = 25
            ws.column_dimensions['E'].width = 30
            
            if cached_data and not cached_data.get('error'):
                cls._add_data_worksheet(wb, context.report_type, cached_data, header_font, header_fill, border)
            
            wb.save(buffer)
            buffer.seek(0)
            return buffer
            
        except Exception as e:
            logger.error(f"Error generating unified Excel: {e}")
            return None
    
    @classmethod
    def _add_report_data_to_excel(cls, ws, row: int, report_type: UnifiedReportType, data: Dict, header_font, header_fill, border) -> int:
        """Add report-type specific data to main Excel sheet"""
        if not data or data.get('error'):
            return row
        
        dog_info = data.get('dog', {})
        if dog_info:
            ws[f'A{row}'] = 'معلومات الكلب'
            ws[f'A{row}'].font = Font(bold=True, size=11, color='3A6EA5')
            row += 1
            
            if dog_info.get('name'):
                ws[f'A{row}'] = 'اسم الكلب'
                ws[f'B{row}'] = dog_info.get('name')
                ws[f'A{row}'].border = border
                ws[f'B{row}'].border = border
                row += 1
            if dog_info.get('code'):
                ws[f'A{row}'] = 'رمز الكلب'
                ws[f'B{row}'] = dog_info.get('code')
                ws[f'A{row}'].border = border
                ws[f'B{row}'].border = border
                row += 1
            row += 1
        
        if data.get('notes'):
            ws[f'A{row}'] = 'ملاحظات'
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
            ws[f'A{row}'] = data.get('notes')
            ws.merge_cells(f'A{row}:D{row}')
            row += 1
        
        return row
    
    @classmethod
    def _add_data_worksheet(cls, wb: Workbook, report_type: UnifiedReportType, data: Dict, header_font, header_fill, border):
        """Add a data worksheet with tables for list data"""
        training = data.get('training_sessions', [])
        if training:
            ws = wb.create_sheet('جلسات التدريب')
            ws.right_to_left = True
            
            headers = ['نوع التدريب', 'الوصف', 'من', 'إلى', 'ملاحظات']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border
                cell.alignment = Alignment(horizontal='right')
            
            for row_idx, session in enumerate(training, 2):
                values = [
                    session.get('training_type', ''),
                    session.get('description', ''),
                    session.get('time_from', '') or '',
                    session.get('time_to', '') or '',
                    session.get('notes', '') or ''
                ]
                for col, val in enumerate(values, 1):
                    cell = ws.cell(row=row_idx, column=col, value=val)
                    cell.border = border
                    cell.alignment = Alignment(horizontal='right')
            
            ws.auto_filter.ref = f"A1:E{len(training)+1}"
            for col_idx in range(1, 6):
                ws.column_dimensions[get_column_letter(col_idx)].width = 20
        
        incidents = data.get('incidents', [])
        if incidents:
            ws = wb.create_sheet('الحوادث')
            ws.right_to_left = True
            
            headers = ['النوع', 'الوصف', 'التاريخ/الوقت', 'الموقع']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border
                cell.alignment = Alignment(horizontal='right')
            
            for row_idx, incident in enumerate(incidents, 2):
                values = [
                    incident.get('incident_type', ''),
                    incident.get('description', ''),
                    incident.get('incident_datetime', '') or '',
                    incident.get('location', '') or ''
                ]
                for col, val in enumerate(values, 1):
                    cell = ws.cell(row=row_idx, column=col, value=val)
                    cell.border = border
                    cell.alignment = Alignment(horizontal='right')
            
            ws.auto_filter.ref = f"A1:D{len(incidents)+1}"
            for col_idx in range(1, 5):
                ws.column_dimensions[get_column_letter(col_idx)].width = 22
        
        health = data.get('health')
        if health:
            ws = wb.create_sheet('الفحص الصحي')
            ws.right_to_left = True
            
            headers = ['الجزء', 'الحالة', 'الملاحظات']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border
                cell.alignment = Alignment(horizontal='right')
            
            health_fields = [
                ('eyes_status', 'eyes_notes', 'العيون'),
                ('nose_status', 'nose_notes', 'الأنف'),
                ('ears_status', 'ears_notes', 'الأذنين'),
                ('mouth_status', 'mouth_notes', 'الفم'),
            ]
            
            row_idx = 2
            for status_field, notes_field, label in health_fields:
                status = health.get(status_field)
                if status:
                    ws.cell(row=row_idx, column=1, value=label).border = border
                    ws.cell(row=row_idx, column=2, value=cls._get_health_status_ar(status)).border = border
                    ws.cell(row=row_idx, column=3, value=health.get(notes_field, '')).border = border
                    for col in range(1, 4):
                        ws.cell(row=row_idx, column=col).alignment = Alignment(horizontal='right')
                    row_idx += 1
            
            for col_idx in range(1, 4):
                ws.column_dimensions[get_column_letter(col_idx)].width = 25
    
    @classmethod
    def _log_export(
        cls,
        context_id: str,
        user_id: str,
        export_format: ExportFormat,
        success: bool,
        error_message: Optional[str] = None
    ):
        """Log an export action"""
        try:
            ip_address = None
            try:
                ip_address = request.remote_addr
            except:
                pass
            
            export_log = ReportExportHistory(
                context_id=context_id,
                export_format=export_format,
                exported_by_user_id=user_id,
                success=success,
                error_message=error_message,
                ip_address=ip_address
            )
            db.session.add(export_log)
            db.session.commit()
        except Exception as e:
            logger.error(f"Error logging export: {e}")
    
    @classmethod
    def _log_approval_action(
        cls,
        context_id: str,
        from_status: Optional[UnifiedReportStatus],
        to_status: UnifiedReportStatus,
        action: str,
        action_by_user_id: str,
        notes: Optional[str] = None,
        rejection_reason: Optional[str] = None
    ):
        """Log an approval workflow action"""
        try:
            ip_address = None
            try:
                ip_address = request.remote_addr
            except:
                pass
            
            approval_log = ReportApprovalHistory(
                context_id=context_id,
                from_status=from_status,
                to_status=to_status,
                action=action,
                action_by_user_id=action_by_user_id,
                notes=notes,
                rejection_reason=rejection_reason,
                ip_address=ip_address
            )
            db.session.add(approval_log)
        except Exception as e:
            logger.error(f"Error logging approval action: {e}")
    
    @classmethod
    def get_pending_reports_for_pm(cls, pm_user_id: str, project_id: Optional[str] = None) -> List[ReportContext]:
        """Get all submitted reports pending PM review"""
        query = ReportContext.query.filter_by(status=UnifiedReportStatus.SUBMITTED)
        
        if project_id:
            query = query.filter_by(project_id=project_id)
        
        return query.order_by(desc(ReportContext.submitted_at)).all()
    
    @classmethod
    def get_pending_reports_for_admin(cls, admin_user_id: str) -> List[ReportContext]:
        """Get all reports pending admin review"""
        if not PermissionService.has_any_permission(admin_user_id, ['admin.*', 'admin.reports.approve']):
            return []
        
        return ReportContext.query.filter(
            ReportContext.status.in_([
                UnifiedReportStatus.PM_APPROVED,
                UnifiedReportStatus.FORWARDED_TO_ADMIN
            ])
        ).order_by(desc(ReportContext.pm_reviewed_at)).all()
    
    @classmethod
    def get_report_history(cls, context_id: str) -> List[ReportApprovalHistory]:
        """Get approval history for a report"""
        return ReportApprovalHistory.query.filter_by(
            context_id=context_id
        ).order_by(desc(ReportApprovalHistory.action_at)).all()
    
    @classmethod
    def get_export_history(cls, context_id: str) -> List[ReportExportHistory]:
        """Get export history for a report"""
        return ReportExportHistory.query.filter_by(
            context_id=context_id
        ).order_by(desc(ReportExportHistory.exported_at)).all()
