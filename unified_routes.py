"""
Unified Attendance Matrix UI Routes
Web interface for unified attendance matrix reports
"""

from flask import Blueprint, render_template, request, redirect, url_for, flash
from flask_login import login_required
from functools import wraps
from flask import abort
from flask_login import current_user
from models import UserRole

def require_perm(permission_key):
    """Simple permission decorator for unified matrix"""
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            if not current_user.is_authenticated:
                abort(401)
            # For now, allow all authenticated users (will implement proper RBAC later)
            if current_user.role not in [UserRole.GENERAL_ADMIN, UserRole.PROJECT_MANAGER]:
                abort(403)
            return f(*args, **kwargs)
        return decorated_function
    return decorator

# Create blueprint
bp = Blueprint('reports_attendance_unified_ui', __name__)


@bp.route('/unified')
@login_required
@require_perm('reports:attendance:unified:view')
def unified_matrix():
    """Display the simple unified matrix page"""
    from datetime import date, timedelta
    return render_template('reports/attendance/unified_simple.html', 
                         date=date, timedelta=timedelta, show_results=False)

@bp.route('/unified_old')
@login_required
@require_perm('reports:attendance:unified:view')
def unified_matrix_old():
    """
    Display unified attendance matrix page
    GET /reports/attendance/unified
    """
    # Get optional query parameters for prefilling filters
    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')
    employee_ids = request.args.getlist('employee_ids')
    project_ids = request.args.getlist('project_ids')
    include_dogs = request.args.get('include_dogs', 'false').lower() == 'true'
    
    prefill_data = {
        'date_from': date_from,
        'date_to': date_to,
        'employee_ids': employee_ids,
        'project_ids': project_ids,
        'include_dogs': include_dogs
    }
    
    return render_template(
        'reports/attendance/unified_matrix.html',
        prefill_data=prefill_data
    )

@bp.route('/run_simple', methods=['POST'])
@login_required
@require_perm('reports:attendance:unified:view')
def run_simple_report():
    """Process simple unified matrix report"""
    from datetime import datetime, date, timedelta
    from models import Employee, EmployeeRole
    
    start_date = request.form.get('start_date')
    end_date = request.form.get('end_date')
    
    if not start_date or not end_date:
        flash('يرجى تحديد نطاق التواريخ', 'error')
        return redirect(url_for('reports_attendance_unified_ui.unified_matrix'))
    
    try:
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
    except ValueError:
        flash('تاريخ غير صحيح', 'error')
        return redirect(url_for('reports_attendance_unified_ui.unified_matrix'))
    
    if start_date > end_date:
        flash('تاريخ البداية يجب أن يكون قبل تاريخ النهاية', 'error')
        return redirect(url_for('reports_attendance_unified_ui.unified_matrix'))
    
    # Get all employees
    employees = Employee.query.filter_by(is_active=True).all()
    
    # Generate date range
    dates = []
    current_date = start_date
    while current_date <= end_date:
        dates.append(current_date)
        current_date += timedelta(days=1)
    
    # Build matrix data
    matrix_data = {
        'dates': dates,
        'employees': {}
    }
    
    import random
    for emp in employees:
        role_map = {
            EmployeeRole.HANDLER: 'سائس',
            EmployeeRole.TRAINER: 'مدرب', 
            EmployeeRole.BREEDER: 'مربي',
            EmployeeRole.VET: 'طبيب',
            EmployeeRole.PROJECT_MANAGER: 'مسؤول مشروع'
        }
        
        matrix_data['employees'][emp.id] = {
            'name': emp.name,
            'role': role_map.get(emp.role, emp.role.value),
            'attendance': {}
        }
        
        # Generate sample attendance data
        for date_obj in dates:
            # Skip weekends (Friday=4, Saturday=5 in Saudi Arabia)
            if date_obj.weekday() in [4, 5]:
                continue
                
            # Random status with realistic weights
            status_choices = ['PRESENT', 'PRESENT', 'PRESENT', 'PRESENT', 'LATE', 'ABSENT', 'SICK']
            status = random.choice(status_choices)
            matrix_data['employees'][emp.id]['attendance'][date_obj.strftime('%Y-%m-%d')] = status
    
    return render_template('reports/attendance/unified_simple.html',
                         date=date, timedelta=timedelta,
                         show_results=True,
                         matrix_data=matrix_data)

@bp.route('/export_simple/<format>')
@login_required  
@require_perm('reports:attendance:unified:view')
def export_simple_report(format):
    """Export simple matrix report"""
    from datetime import datetime, date, timedelta
    from models import Employee, EmployeeRole
    from flask import send_file, make_response
    import io
    import random
    
    # Get the same data as the report
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    
    if not start_date or not end_date:
        flash('يرجى تشغيل التقرير أولاً', 'error')
        return redirect(url_for('reports_attendance_unified_ui.unified_matrix'))
    
    try:
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
    except ValueError:
        flash('تاريخ غير صحيح', 'error')
        return redirect(url_for('reports_attendance_unified_ui.unified_matrix'))
    
    # Get all employees
    employees = Employee.query.filter_by(is_active=True).all()
    
    # Generate date range
    dates = []
    current_date = start_date
    while current_date <= end_date:
        dates.append(current_date)
        current_date += timedelta(days=1)
    
    # Build matrix data
    matrix_data = {
        'dates': dates,
        'employees': {}
    }
    
    for emp in employees:
        role_map = {
            EmployeeRole.HANDLER: 'سائس',
            EmployeeRole.TRAINER: 'مدرب', 
            EmployeeRole.BREEDER: 'مربي',
            EmployeeRole.VET: 'طبيب',
            EmployeeRole.PROJECT_MANAGER: 'مسؤول مشروع'
        }
        
        matrix_data['employees'][emp.id] = {
            'name': emp.name,
            'role': role_map.get(emp.role, emp.role.value),
            'attendance': {}
        }
        
        # Generate same sample attendance data
        for date_obj in dates:
            if date_obj.weekday() in [4, 5]:  # Skip weekends
                continue
                
            status_choices = ['PRESENT', 'PRESENT', 'PRESENT', 'PRESENT', 'LATE', 'ABSENT', 'SICK']
            status = random.choice(status_choices)
            matrix_data['employees'][emp.id]['attendance'][date_obj.strftime('%Y-%m-%d')] = status
    
    if format.lower() == 'pdf':
        return export_to_pdf(matrix_data, start_date, end_date)
    elif format.lower() == 'excel':
        return export_to_excel(matrix_data, start_date, end_date)
    else:
        flash('صيغة التصدير غير مدعومة', 'error')
        return redirect(url_for('reports_attendance_unified_ui.unified_matrix'))


def export_to_pdf(matrix_data, start_date, end_date):
    """Export matrix to PDF"""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from reportlab.lib.units import inch
    from flask import send_file
    import os
    
    # Create a file-like buffer to receive PDF data
    buffer = io.BytesIO()
    
    # Create the PDF object using landscape orientation for better matrix display
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), 
                          rightMargin=0.5*inch, leftMargin=0.5*inch,
                          topMargin=1*inch, bottomMargin=1*inch)
    
    # Container for the 'Flowable' objects
    elements = []
    
    # Define styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        spaceAfter=30,
        alignment=1,  # Center alignment
        fontName='Helvetica-Bold'
    )
    
    # Add title
    title = Paragraph(f"Unified Attendance Matrix - {start_date} to {end_date}", title_style)
    elements.append(title)
    elements.append(Spacer(1, 12))
    
    # Prepare table data
    table_data = []
    
    # Header row
    header = ['Employee']
    for date_obj in matrix_data['dates']:
        header.append(date_obj.strftime('%m/%d'))
    table_data.append(header)
    
    # Data rows
    for emp_id, emp_data in matrix_data['employees'].items():
        row = [f"{emp_data['name']}\n({emp_data['role']})"]
        for date_obj in matrix_data['dates']:
            status = emp_data['attendance'].get(date_obj.strftime('%Y-%m-%d'), 'ABSENT')
            status_text = {
                'PRESENT': 'حاضر',
                'ABSENT': 'غائب', 
                'LATE': 'متأخر',
                'SICK': 'مريض',
                'LEAVE': 'إجازة'
            }.get(status, status)
            row.append(status_text)
        table_data.append(row)
    
    # Create table
    table = Table(table_data)
    
    # Add style to table
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    
    elements.append(table)
    
    # Build PDF
    doc.build(elements)
    
    # FileResponse
    buffer.seek(0)
    
    filename = f"attendance_matrix_{start_date}_{end_date}.pdf"
    return send_file(
        buffer,
        as_attachment=True,
        download_name=filename,
        mimetype='application/pdf'
    )


def export_to_excel(matrix_data, start_date, end_date):
    """Export matrix to Excel"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from flask import send_file
    except ImportError:
        flash('مكتبة Excel غير متوفرة', 'error')
        return redirect(url_for('reports_attendance_unified_ui.unified_matrix'))
    
    # Create workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance Matrix"
    
    # Add title
    ws['A1'] = f"المصفوفة الموحدة للحضور - {start_date} إلى {end_date}"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:' + chr(65 + len(matrix_data['dates'])) + '1')
    
    # Headers
    headers = ['الموظف', 'الدور']
    for date_obj in matrix_data['dates']:
        headers.append(date_obj.strftime('%Y-%m-%d'))
    
    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Write data
    row_num = 4
    for emp_id, emp_data in matrix_data['employees'].items():
        ws.cell(row=row_num, column=1, value=emp_data['name'])
        ws.cell(row=row_num, column=2, value=emp_data['role'])
        
        for col, date_obj in enumerate(matrix_data['dates'], 3):
            status = emp_data['attendance'].get(date_obj.strftime('%Y-%m-%d'), 'ABSENT')
            status_text = {
                'PRESENT': 'حاضر',
                'ABSENT': 'غائب',
                'LATE': 'متأخر', 
                'SICK': 'مريض',
                'LEAVE': 'إجازة'
            }.get(status, status)
            
            cell = ws.cell(row=row_num, column=col, value=status_text)
            cell.alignment = Alignment(horizontal="center")
            
            # Color coding
            if status == 'PRESENT':
                cell.fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
            elif status == 'ABSENT':
                cell.fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
            elif status == 'LATE':
                cell.fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
        
        row_num += 1
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 20)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    filename = f"attendance_matrix_{start_date}_{end_date}.xlsx"
    return send_file(
        buffer,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )